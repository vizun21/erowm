# -*- coding: utf-8 -*-

from django.shortcuts import render,redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from .models import Owner, Business, Profile, Sales, Agency, Account
from .models import Subsection, Paragraph, Item, Subdivision
from .models import Transaction, TBLBANK, Budget, Deadline
from .forms import SignupForm, OwnerForm, BusinessForm, UserForm, SalesForm, AgencyForm, EditOwnerForm, AccountForm
from .forms import SubsectionForm, ParagraphForm, ItemForm, SubdivisionForm
from .forms import TblbankDirectForm
from .crypto import AESCipher

from django.contrib.auth import authenticate, login
from django.http import HttpResponse, HttpResponseRedirect
import datetime
from django.utils import timezone
from django.utils.dateformat import DateFormat
from dateutil.relativedelta import relativedelta

from django.db.models import Sum, Count, Case, When
from django.db.models.functions import Coalesce
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

import json

# Create your views here.

ACCOUNTANT = 1
OWNER = 2
SALES = 3
AGENCY = 4
LOCAL = 5

def main(request):
    user = request.user
    if not user.is_active:
        return redirect("login")
    else:
        request.session['master_login'] = False
        if user.profile.level_id >= LOCAL:
            return redirect("user_list")
        elif user.profile.level_id >= SALES:
            return redirect("user_transform")
        else:
            return redirect("business_list")
    return render(request, 'accounting/main.html')

def id_check(request):
    id_check = False
    if request.method == "POST":
        signupform = SignupForm(request.POST)
        ownerform = OwnerForm(request.POST)
        if signupform.is_valid() and ownerform.is_valid():
            user = signupform.save(commit=False)
            user.email = signupform.cleaned_data['email']
            user.save()
            Profile.objects.create(user=user, level_id=OWNER)
            profile =  Profile.objects.get(user=user)
            owner = ownerform.save(commit=False)
            owner.profile = profile
            owner.email = signupform.cleaned_data['email']
            owner.save()
            password = user.password[34:]
            id_check = True
            return HttpResponse(json.dumps({'id_check': id_check, 'password': password}), content_type="application/json")
    return HttpResponse(json.dumps({'id_check': id_check}), content_type="application/json")

def signup(request):
    if request.method == "POST":
        signup_done = request.POST.get('signup_done')
        if signup_done:
            return redirect("signup_done")
        else:
            signupform = SignupForm(request.POST)
            ownerform = OwnerForm(request.POST)

    elif request.method == "GET":
        signupform = SignupForm()
        ownerform = OwnerForm()

    return render(request, "registration/signup.html", {
        "signupform": signupform,
        "ownerform": ownerform,
    })

def signup_done(request):
    return render(request, "registration/signup_done.html")

@login_required(login_url='/')
def user_delete(request):
    if request.user.profile.level_id < LOCAL:
        return HttpResponse("<script>alert('권한이 없습니다.');history.back();</script>")
    username = request.POST.get('username')
    user = get_object_or_404(User, username=username)
    user.delete()
    return redirect('user_list')

@login_required(login_url='/')
def business_list(request):
    owner = request.user.profile.owner
    lists = Business.objects.filter(owner=owner)
    return render(request, 'accounting/business_list.html', {'lists': lists, 'master_login': request.session['master_login']})

@login_required(login_url='/')
def business_create(request):
    owner = request.user.profile.owner
    if request.method == "POST":
        form = BusinessForm(request.POST)
        if form.is_valid():
            business = form.save(commit=False)
            business.owner = owner
            business.save()
            return redirect('business_list')
    else:
        form = BusinessForm(initial={'place_name': owner.place_name, 'reg_number': owner.reg_number, 'owner_name': owner.name, 'owner_reg_number1': owner.owner_reg_number1, 'owner_reg_number2': owner.owner_reg_number2, 'type1': owner.type1, 'type2': owner.type2, 'cellphone': owner.cellphone, 'phone': owner.phone, 'fax': owner.fax, 'email': owner.email, 'zip_number': owner.zip_number, 'address': owner.address, 'detailed_address': owner.detailed_address})
    return render(request, 'accounting/business_edit.html', {'form': form})

@login_required(login_url='/')
def business_edit(request, pk):
    business = get_object_or_404(Business, pk=pk)
    owner = request.user.profile.owner
    if request.method == "POST":
        form = BusinessForm(request.POST, instance=business)
        if form.is_valid():
            business = form.save(commit=False)
            business.owner = owner
            business.save()
            return redirect('business_list')
    else:
        form = BusinessForm(instance=business)
    return render(request, 'accounting/business_edit.html', {'form': form, 'business': business})

@login_required(login_url='/')
def transform_business(request, pk):
    business = get_object_or_404(Business, pk=pk)
    request.session['business'] = business.pk
    return redirect('business_info')

@login_required(login_url='/')
def retransform_business(request):
    del request.session['business']
    return redirect('business_list')

@login_required(login_url='/')
def business_info(request):
    owner = request.user.profile.owner
    business = get_object_or_404(Business, pk=request.session['business'])
    if business.owner != owner:
        return redirect('business_list')
    return render(request, 'accounting/business_info.html', {
        'business': business, 'business_info_page': 'active', 'business_management': 'active', 'master_login': request.session['master_login']
    })

@login_required(login_url='/')
def account_list(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    lists = Account.objects.filter(business=business)
    return render(request, 'accounting/account_list.html', {'lists': lists, 'business': business, 'accounting_management': 'active', 'account_list': 'active', 'master_login': request.session['master_login']})

@login_required(login_url='/')
def account_create(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    main_acct = Account.objects.filter(business=business, main=True).count()
    if request.method == "POST":
        regist_done = request.POST.get('regist_done')
        if regist_done:
            return redirect("account_list")
        else:
            form = AccountForm(request.POST)
            if form.is_valid():
                account = form.save(commit=False)
                if main_acct:
                    account.main = False
                account.save()
                return redirect('account_list')
    else:
        form = AccountForm(initial={'business': business})
    return render(request, 'accounting/account_create.html', {'form': form, 'business': business})

@login_required(login_url='/')
def account_edit(request, pk):
    account = get_object_or_404(Account, pk=pk)
    business = get_object_or_404(Business, pk=request.session['business'])
    main_acct = Account.objects.filter(business=business, main=True).count()
    if request.method == "POST":
        regist_done = request.POST.get('regist_done')
        if regist_done:
            return redirect("account_list")
        else:
            form = AccountForm(request.POST, instance=account)
            if form.is_valid():
                account = form.save(commit=False)
                if main_acct:
                    account.main = False
                account.save()
                return redirect('account_list')
    else:
        form = AccountForm(instance=account)
    return render(request, 'accounting/account_edit.html', {'acctpk':pk, 'form': form, 'business': business})

@login_required(login_url='/')
def account_check(request):
    command = request.POST.get("command", "create")
    print(command)
    business = get_object_or_404(Business, pk=request.session['business'])
    account_check = False
    main_acct = Account.objects.filter(business=business, main=True).count()
    if request.method == "POST":
        if command == "edit":
            account = get_object_or_404(Account, pk=request.POST.get("acctpk"))
            form = AccountForm(request.POST, instance=account)
            if form.is_valid():
                account = form.save(commit=False)
                if main_acct:
                    account.main = False
                account.save()
                account_check = True
                password = request.user.password[34:]
                Bjumin = business.reg_number.split("-")
                return HttpResponse(json.dumps({'account_check': account_check, 'username': request.user.username, 'password': password, 'Mjumin_1': business.owner_reg_number1, 'Bjumin_1': Bjumin[0], 'Bjumin_2': Bjumin[1], 'Bjumin_3': Bjumin[2]}), content_type="application/json")
        else:
            accountform = AccountForm(request.POST)
            if accountform.is_valid():
                account = accountform.save(commit=False)
                if main_acct:
                    account.main = False
                account.save()
                account_check = True
                password = request.user.password[34:]
                Bjumin = business.reg_number.split("-")
                return HttpResponse(json.dumps({'account_check': account_check, 'username': request.user.username, 'password': password, 'Mjumin_1': business.owner_reg_number1, 'Bjumin_1': Bjumin[0], 'Bjumin_2': Bjumin[1], 'Bjumin_3': Bjumin[2]}), content_type="application/json")
    return HttpResponse(json.dumps({'account_check': account_check}), content_type="application/json")

@login_required(login_url='/')
def account_delete(request, pk):
    #아이디에해당하는사업장만 삭제할수 있도록
    account = get_object_or_404(Account, pk=pk)
    if request.user.profile.level_id < LOCAL:
    #    if account.business.owner.profile.user != request.user: #사용자페이지삭제구현시 확인필요
        return HttpResponse("<script>alert('권한이 없습니다.');history.back();</script>")
    account.delete()
    return redirect('bankda_account')

@login_required(login_url='/')
def user_list(request):
    if request.user.is_staff:
        owners = Owner.objects.all()
        where = request.GET.get('where', '')
        keyword = request.GET.get('keyword', '')
        sales = Sales.objects.all()
        if where == 'place_name'  and keyword:
            owners = owners.filter(place_name__icontains=keyword)
        elif where == 'name' and keyword:
            owners = owners.filter(name__icontains=keyword)
        elif where == 'sales' and keyword:
            owners = owners.filter(sales__name__icontains=keyword)
    else:
        return redirect('business_list')
    return render(request, 'accounting/user_list.html', {
        'owners': owners, 'sales': sales, 'user_list_page': 'active'
    })

@login_required(login_url='/')
def user_transform(request):
    user = request.user
    if user.is_staff:
        owners = Owner.objects.all()
    elif user.profile.level_id == SALES:
        owners = Owner.objects.filter(sales=user.profile.sales)
    elif user.profile.level_id == AGENCY:
        sales_list = list()
        sales_set = Sales.objects.filter(agency=user.profile.agency)
        for sales in sales_set:
            sales_list.append(sales.id)
        owners = Owner.objects.filter(sales_id__in=sales_list)
    else:
        return redirect('business_list')

    where = request.GET.get('where', '')
    keyword = request.GET.get('keyword', '')
    if where == 'place_name'  and keyword:
        owners = owners.filter(place_name__icontains=keyword)
    elif where == 'name' and keyword:
        owners = owners.filter(name__icontains=keyword)
    
    owner_exists = owners.exists()
    
    return render(request, 'accounting/user_transform.html', {
        'owners': owners, 'owner_exists': owner_exists, 'user_transform_page': 'active'})

@login_required(login_url='/')
def transform(request, pk):
    user = get_object_or_404(User, pk=pk)
    if (request.user.profile.level_id >= SALES) and user.is_active:
        master_username = request.user.username
        login(request, user)
        if user is not None:
            request.session['master_login'] = True
            request.session['username'] = master_username
            return redirect('business_list')
    return redirect('user_transform')

@login_required(login_url='/')
def retransform(request):
    if request.session['master_login'] == True:
        username = request.session['username']
        user = User.objects.get(username=username)
        login(request, user)
        if user is not None:
            request.session['username'] = user.username
            request.session['master_login'] = False
            return redirect('user_transform')
    return redirect('business_list')

@login_required(login_url='/')
def sales_list(request):
    user = request.user
    if user.is_staff:
        sales = Sales.objects.all()
        agency = Agency.objects.all()
        where = request.GET.get('where', '')
        keyword = request.GET.get('keyword', '')
        if where == 'name'  and keyword:
            sales = sales.filter(name__icontains=keyword)
        elif where == 'agency' and keyword:
            sales = sales.filter(agency__name__contains=keyword)
    else:
        return redirect('business_list')
    return render(request, 'accounting/sales_list.html', {
        'sales': sales, 'agency': agency, 'sales_list_page': 'active'
    })

@login_required(login_url='/')
def sales_create(request):
    if request.method == "POST":
        userform = UserForm(request.POST)
        form = SalesForm(request.POST)
        if userform.is_valid() and form.is_valid():
            user = userform.save(commit=False)
            user.email = userform.cleaned_data['email']
            user.save()
            Profile.objects.create(user=user, level_id=SALES)
            profile = Profile.objects.get(user=user)
            sales = form.save(commit=False)
            sales.profile = profile
            sales.email = userform.cleaned_data['email']
            sales.save()
            return redirect('sales_list')
    else:
        userform = UserForm()
        form = SalesForm()
    return render(request, 'accounting/sales_edit.html', {'form': form, 'userform': userform, 'sales_list_page': 'active'})

@login_required(login_url='/')
def sales_edit(request, pk):
    sales = get_object_or_404(Sales, pk=pk)
    if request.method == "POST":
        form = SalesForm(request.POST, instance=sales)
        if form.is_valid():
            form.save()
            return redirect('sales_list')
    else:
        form = SalesForm(instance=sales)
    return render(request, 'accounting/sales_edit.html', {'form': form, 'sales_list_page': 'active'})

@login_required(login_url='/')
def sales_delete(request, pk):
    sales = get_object_or_404(Sales, pk=pk)
    sales.delete()
    return redirect('sales_list')

@login_required(login_url='/')
def sales_change(request):
    pk_list = request.POST.getlist('check_list[]')
    if request.method == "POST":
        auth_selected = request.POST.get('select_auth', '')
        sales_selected = request.POST.get('select_sales', '')
        for pk in pk_list:
            owner = get_object_or_404(Owner, pk=pk)
            owner.is_demo = auth_selected
            owner.sales_id = sales_selected
            owner.save()
    return redirect('user_list')

@login_required(login_url='/')
def agency_list(request):
    user = request.user
    if user.is_staff:
        agency = Agency.objects.all()
        where = request.GET.get('where', '')
        keyword = request.GET.get('keyword', '')
        if where == 'name'  and keyword:
            agency = agency.filter(name__icontains=keyword)
        elif where == 'owner_name' and keyword:
            agency = agency.filter(owner_name__icontains=keyword)
    else:
        return redirect('business_list')
    return render(request, 'accounting/agency_list.html', {
        'agency': agency, 'agency_list_page': 'active'
    })

@login_required(login_url='/')
def agency_create(request):
    if request.method == "POST":
        userform = UserForm(request.POST)
        form = AgencyForm(request.POST)
        if userform.is_valid() and form.is_valid():
            user = userform.save(commit=False)
            user.email = userform.cleaned_data['email']
            user.save()
            Profile.objects.create(user=user, level_id=AGENCY)
            profile = Profile.objects.get(user=user)
            agency = form.save(commit=False)
            agency.profile = profile
            agency.email = userform.cleaned_data['email']
            agency.save()
            return redirect('agency_list')
    else:
        userform = UserForm()
        form = AgencyForm()
    return render(request, 'accounting/agency_edit.html', {'form': form, 'userform': userform, 'agency_list_page': 'active'})

@login_required(login_url='/')
def agency_edit(request, pk):
    agency = get_object_or_404(Agency, pk=pk)
    if request.method == "POST":
        form = AgencyForm(request.POST, instance=agency)
        if form.is_valid():
            form.save()
            return redirect('agency_list')
    else:
        form = AgencyForm(instance=agency)
    return render(request, 'accounting/agency_edit.html', {'form': form, 'agency_list_page': 'active'})

@login_required(login_url='/')
def agency_delete(request, pk):
    agency = get_object_or_404(Agency, pk=pk)
    agency.delete()
    return redirect('agency_list')

@login_required(login_url='/')
def agency_change(request):
    pk_list = request.POST.getlist('check_list[]')
    if request.method == "POST":
        selected_agency = request.POST.get('select_agency', '')
        for pk in pk_list:
            sales = get_object_or_404(Sales, pk=pk)
            sales.agency_id = selected_agency
            sales.save()
    return redirect('sales_list')

@login_required(login_url='/')
def mypage(request):
    owner = request.user.profile.owner
    return render(request, 'accounting/mypage.html', {
        'owner': owner, 'mypage': 'active',
    })

@login_required(login_url='/')
def mypage_edit(request):
    owner = request.user.profile.owner
    if request.method == "POST":
        form = EditOwnerForm(request.POST, instance=owner)
        if form.is_valid():
            owner = form.save(commit=False)
            owner.save()
            return redirect('mypage')
    else:
        form = EditOwnerForm(instance=owner)
    return render(request, 'accounting/mypage_edit.html', {'form': form})

@login_required(login_url='/')
def bankda_join(request):
    owners = Owner.objects.all()
    return render(request, 'accounting/bankda_join.html', {
        'owners': owners, 'bankda_page': 'active',
    })

@login_required(login_url='/')
def bankda_account(request):
    accounts = Account.objects.all()
    return render(request, 'accounting/bankda_account.html', {
        'accounts': accounts, 'bankda_account_page': 'active',
    })

@login_required(login_url='/')
def transaction_history(request):
    today = datetime.datetime.now()
    business = get_object_or_404(Business, pk=request.session['business'])
    acct_list = Account.objects.filter(business=business)
    try:
        main_acctid = Account.objects.get(business=business, main=True).id
    except:
        return HttpResponse("<script>alert('등록된 계좌가 없습니다. 계좌를 먼저 등록해주세요.');window.location.href='/account/list/'</script>")

    acctid = request.GET.get('acctid', main_acctid)
    year = request.GET.get('year', DateFormat(today).format("Y"))
    month = request.GET.get('month', DateFormat(today).format("m"))
    page = request.GET.get('page', 1)
    page2 = request.GET.get('page2', 1)

    acct = get_object_or_404(Account, business=business, id=acctid)
    start_date = datetime.datetime.strptime(year+'-'+month+'-01', '%Y-%m-%d')
    end_date = start_date + relativedelta(months=1)

    input_items = Item.objects.filter(paragraph__subsection__institution = business.type3).filter(paragraph__subsection__type="수입").exclude(paragraph__subsection__code=0)
    output_items = Item.objects.filter(paragraph__subsection__institution = business.type3).filter(paragraph__subsection__type="지출")
    input_subsections = Subsection.objects.filter(institution=business.type3).filter(type="수입")

    selected_check_list = []
    for idx, val in enumerate(request.POST.getlist('tr_check_list')):
        selected_check_list.append(int(val))
    item_list = request.POST.getlist('item_list')
    selected_item_list = []
    subdivision_list = []
    for idx, val in enumerate(item_list):
        if val:
            selected_item_list.append(int(val))
        else:
            selected_item_list.append(0)
        subdivisions = Subdivision.objects.filter(item__id = selected_item_list[idx])
        subdivision_list.append(subdivisions)

    input_subsection_list = request.POST.getlist('input_subsection_list')
    selected_subsection_list = []
    relative_item_list = []
    for idx, val in enumerate(input_subsection_list):
        if val:
            selected_subsection_list.append(int(val))
        else:
            selected_subsection_list.append(0)
        relative_item = Item.objects.filter(paragraph__subsection__id = selected_subsection_list[idx])
        relative_item_list.append(relative_item)

    subdivision_lists = request.POST.getlist('subdivision_list')
    selected_subdivision_list = []
    for idx, val in enumerate(subdivision_lists):
        if val:
            selected_subdivision_list.append(int(val))
        else:
            selected_subdivision_list.append(0)

    input_subdivision_list = request.POST.getlist('input_subdivision_list')
    selected_input_subdivision_list = []
    for idx, val in enumerate(input_subdivision_list):
        if val:
            selected_input_subdivision_list.append(int(val))
        else:
            selected_input_subdivision_list.append(0)

    data_list = TBLBANK.objects.filter(Bkacctno=acct.account_number, Bkdate__gte=start_date, Bkdate__lt=end_date).order_by('-Bkdate', '-Bkid', 'Bkdivision')
    transaction_list = Transaction.objects.filter(business=business, Bkdate__gte=start_date, Bkdate__lt=end_date).order_by('-Bkdate', '-id', 'Bkdivision')
    #print(start_date, data_list[0].Bkdate)

    #--수입/지출계
    try:
        premonth_transfer_price = transaction_list.get(Bkdate=start_date, Bkdivision=0).Bkjango
    except:
        premonth_transfer_price = 0

    total_input = 0
    total_output = 0

    for transaction in transaction_list:
        if transaction.Bkinput != 0:
            total_input += transaction.Bkinput
        elif transaction.Bkoutput != 0:
            total_output += transaction.Bkoutput
    total_input -= premonth_transfer_price

    try:
        jango = transaction_list.first().Bkjango
    except:
        jango = 0

    paginator = Paginator(data_list, 10)
    paginator2 = Paginator(transaction_list, 10)
    
    try:
        data = paginator.page(page)
    except PageNotAnInteger:
        data = paginator.page(1)
    except EmptyPage:
        data = Paginator.page(paginator.num_pages)

    try:
        data2 = paginator2.page(page2)
    except PageNotAnInteger:
        data2 = paginator2.page(1)
    except EmptyPage:
        data2 = Paginator.page(paginator2.num_pages)
    return render(request, 'accounting/transaction_history.html', {
        'selected_check_list': selected_check_list, 'total_input': total_input, 'total_output': total_output, 'jango': jango, 'premonth_transfer_price': premonth_transfer_price, 'year': int(year), 'month': int(month), 'year_range': range(int(DateFormat(today).format("Y")), 1999, -1), 'month_range': range(1,13), 'acctid': int(acctid), 'page': page, 'page2': page2, 'acct_list': acct_list, 'selected_subdivision_list': selected_subdivision_list, 'selected_input_subdivision_list': selected_input_subdivision_list, 'relative_item_list': relative_item_list, 'input_subsections': input_subsections, 'selected_subsection_list': selected_subsection_list, 'selected_item_list': selected_item_list, 'data': data, 'data2': data2, 'input_items': input_items, 'output_items': output_items, 'subdivision_list': subdivision_list, 'business': business, 'accounting_management': 'active', 'transaction_history_page': 'active', 'master_login': request.session['master_login'],
    })

@login_required(login_url='/')
def regist_transaction(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    acctid = request.POST.get('acctid')
    acct = get_object_or_404(Account, business=business, id=acctid)
    Mid = request.user.username
    today = datetime.datetime.now()
    year = request.POST.get('year')
    month = request.POST.get('month')
    page = request.POST.get('page')
    page2 = request.POST.get('page2')
    tr_check_list = request.POST.getlist('tr_check_list')
    tr_list = request.POST.getlist('transaction_list')
    Bkdate_list = request.POST.getlist('Bkdate_list')
    Bkjukyo_list = request.POST.getlist('Bkjukyo_list')
    item_list = request.POST.getlist('item_list')
    subdivision_list = request.POST.getlist('subdivision_list')
    relative_subsection_list = request.POST.getlist('input_subsection_list')
    relative_item_list = request.POST.getlist('input_subdivision_list')

    try:
        close = Deadline.objects.get(business=business, year=year, month=month)
        if close.regdatetime:
            return HttpResponse("<script>alert('해당월은 마감완료되었습니다.');history.back();</script>")
    except:
        pass

    if tr_check_list == []:
        return HttpResponse("<script>alert('선택된 거래가 없습니다.');history.back();</script>")
    for check in tr_check_list:
        if item_list[int(check)] == '':
            return HttpResponse("<script>alert('계정명을 선택해주세요.');history.back();</script>")

    tr_check_list.reverse()

    if request.method == "POST":
        for check in tr_check_list:
            Bkid=tr_list[int(check)]
            tblbank_tr = TBLBANK.objects.get(Bkid=Bkid)
            Bkdivision = Transaction.objects.filter(Bkid=Bkid).filter(Bkdivision__gt=0).count() + 1
            Bkdate = Bkdate_list[int(check)]
            start_date = datetime.datetime.strptime(Bkdate[:8]+'01', "%Y-%m-%d")

            a_month_ago = start_date - relativedelta(months=1)
            a_month_later = start_date + relativedelta(months=1)

            #--해당날짜 전월이월금 유무확인--
            try :
                carryover_tr = Transaction.objects.filter(business=business, Bkdate=start_date).get(Bkdivision=0)
            except Transaction.DoesNotExist:
                #--전월이월금 없는 경우 주계좌의 이전달 마지막 내역을 전월이월금으로 등록
                main_acct = Account.objects.get(business=business, main=True)
                last_tr = TBLBANK.objects.filter(Bkacctno=main_acct.account_number, Bkdate__gte=a_month_ago).filter(Bkdate__lt=start_date).order_by('Bkdate', 'Bkid').last()
                if last_tr == None:
                    return HttpResponse("<script>alert('주계좌의 전월 거래내역이 없습니다. 전월거래내역이 있는 계좌를 주계좌로 변경하세요.');history.back();</script>")
                Transaction.objects.create(
                    Bkid=last_tr.Bkid,
                    Bkdivision=0,
                    Mid=Mid,
                    business=business,
                    Bkacctno=main_acct.account_number,
                    Bkname=main_acct.bank.name,
                    Bkdate=start_date,
                    Bkjukyo="전월이월금",
                    Bkinput=last_tr.Bkjango,
                    Bkoutput=0,
                    Bkjango=last_tr.Bkjango,
                    item=Item.objects.get(paragraph__subsection__institution=business.type3, paragraph__subsection__code=0, paragraph__code=0, code=0),
                    regdatetime=today
                )

            Bkjukyo = Bkjukyo_list[int(check)]
            Bkinput = tblbank_tr.Bkinput
            Bkoutput = tblbank_tr.Bkoutput
            latest_tr = Transaction.objects.filter(business=business, Bkdate__lte=Bkdate).order_by('-Bkdate', '-id').first()
            jango = latest_tr.Bkjango
            Bkjango = 0
            if Bkinput > 0:
                Bkjango = jango + Bkinput
            if Bkoutput > 0:
                Bkjango = jango - Bkoutput
            item = Item.objects.get(id=item_list[int(check)])
            try:
                subdivision = Subdivision.objects.get(id=subdivision_list[int(check)])
            except Subdivision.DoesNotExist:
                subdivision = None
            try:
                relative_subsection = Subsection.objects.get(id=relative_subsection_list[int(check)])
            except Subsection.DoesNotExist:
                relative_subsection = None
            try:
                relative_item = Item.objects.get(id=relative_item_list[int(check)])
            except Item.DoesNotExist:
                relative_item = None

            transaction = Transaction(
                Bkid=Bkid,
                Bkdivision=Bkdivision,
                Mid=Mid,
                business=business,
                Bkacctno=acct.account_number,
                Bkname=acct.bank.name,
                Bkdate=Bkdate,
                Bkjukyo=Bkjukyo,
                Bkinput=Bkinput,
                Bkoutput=Bkoutput,
                Bkjango=Bkjango,
                regdatetime=today,
                item = item,
                subdivision = subdivision,
                relative_subsection = relative_subsection,
                relative_item = relative_item
            )
            transaction.save()

            update_list = Transaction.objects.filter(business=business, Bkdate__gt=Bkdate, Bkdate__lt=a_month_later)
            for update in update_list:
                if transaction.Bkinput > 0:
                    update.Bkjango = update.Bkjango + transaction.Bkinput
                elif transaction.Bkoutput > 0:
                    update.Bkjango = update.Bkjango - transaction.Bkoutput
                update.save()

            tr = TBLBANK.objects.get(Bkid=Bkid)
            tr.sub_Bkjukyo=Bkjukyo
            tr.item = item
            tr.subdivision = subdivision
            tr.relative_subsection = relative_subsection
            tr.relative_item = relative_item
            tr.regdatetime=today
            tr.save()

    response = redirect('transaction_history')
    response['Location'] += '?page='+page+'&page2='+page2+'&year='+year+'&month='+month+'&acctid='+acctid
    return response

@login_required(login_url='/')
def transaction_delete(request):
    Mid = request.user.username
    business = get_object_or_404(Business, pk=request.session['business'])
    check_list = request.POST.getlist('check_list[]')
    acctid = request.POST.get('acctid')
    year = request.POST.get('year')
    month = request.POST.get('month')
    page = request.POST.get('page')
    page2 = request.POST.get('page2')
    
    try:
        close = Deadline.objects.get(business=business, year=year, month=month)
        if close.regdatetime:
            return HttpResponse("<script>alert('해당월은 마감완료되었습니다.');history.back();</script>")
    except:
        pass

    for Bkid in check_list:
        transaction_list = Transaction.objects.filter(Bkid=Bkid).exclude(Bkdivision=0)
        tblbank_list = TBLBANK.objects.filter(Bkid=Bkid).exclude(Bkdivision=0)
        for tr in transaction_list:
            #잔고계산
            a_month_later = DateFormat(tr.Bkdate + relativedelta(months=1)).format("Y-m-01")
            update_list = Transaction.objects.filter(business=business, Bkdate__gte=tr.Bkdate, Bkdate__lt=a_month_later)
            for update in update_list:
                if (update.Bkdate == tr.Bkdate and update.id > tr.id) or update.Bkdate != tr.Bkdate:
                    if tr.Bkinput != 0:
                        update.Bkjango = update.Bkjango - tr.Bkinput
                    elif tr.Bkoutput != 0:
                        update.Bkjango = update.Bkjango + tr.Bkoutput
                    update.save()
            tr.delete()
        for tblbank in tblbank_list:
            if tblbank.Bkdivision == 1 and tblbank.direct == False:
                tblbank.regdatetime = None
                tblbank.item = None
                tblbank.relative_item = None
                tblbank.relative_subsection = None
                tblbank.subdivision = None
                tblbank.sub_Bkjukyo = None
                tblbank.save()
            else:
                tblbank.delete()

    response = redirect('transaction_history')
    response['Location'] += '?page='+page+'&page2='+page2+'&year='+year+'&month='+month+'&acctid='+acctid
    return response

@login_required(login_url='/')
def spi_list(request):
    subsection_list = Subsection.objects.all().annotate(count=Count('paragraph__item')).exclude(count=0).order_by('institution','type','code')
    paragraph_list = Paragraph.objects.all().annotate(count=Count('item')).exclude(count=0).order_by('code')
    item_list = Item.objects.all().order_by('code')
    return render(request, 'accounting/spi_list.html', {'setting_page': "active", 'spi_list_page': "active", 'subsection_list': subsection_list, 'paragraph_list': paragraph_list, 'item_list': item_list})

@login_required(login_url='/')
def subsection_create(request):
    if request.method == "POST":
        form = SubsectionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('spi_list')
    else:
        form = SubsectionForm()
    return render(request, 'accounting/subsection_edit.html', {'form': form})

@login_required(login_url='/')
def paragraph_create(request):
    if request.method == "POST":
        form = ParagraphForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('spi_list')
    else:
        form = ParagraphForm()
    return render(request, 'accounting/paragraph_edit.html', {'form': form})

@login_required(login_url='/')
def item_create(request):
    if request.method == "POST":
        form = ItemForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('spi_list')
    else:
        form = ItemForm()
    return render(request, 'accounting/item_edit.html', {'form': form})

@login_required(login_url='/')
def subdivision_list(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    subdivisions = Subdivision.objects.filter(business=business).order_by('item')
    return render(request, 'accounting/subdivision_list.html', {'business': business, 'subdivisions': subdivisions, 'master_login': request.session['master_login'], 'accounting_management': 'active', 'subdivision_list': 'active'})

@login_required(login_url='/')
def subdivision_create(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    if request.method == "POST":
        form = SubdivisionForm(request.POST)
        if form.is_valid():
            subdivision = form.save(commit=False)
            subdivision.business = business
            subdivision.save()
            return redirect('subdivision_list')
    else:
        form = SubdivisionForm()
        form.fields['item'].queryset = Item.objects.filter(paragraph__subsection__institution = business.type3)
    return render(request, 'accounting/subdivision_edit.html', {'business': business, 'form': form, 'accounting_management': 'active', 'subdivision_list': 'active'})

@login_required(login_url='/')
def other_settings(request):
    return render(request, 'accounting/other_settings.html')

@login_required(login_url='/')
def database_syn(request):
    from .models import Setting
    tr_num = Setting.objects.get(name="transaction_num")

    from django.db import connections
    query = """SELECT * FROM TBLBANK WHERE Bkid > """+str(tr_num.value)

    with connections['default'].cursor() as cursor:
        cursor.execute(query)
        columns = [ col[0] for col in cursor.description ]
        data_list = [ dict(zip(columns,row)) for row in cursor.fetchall() ]

    try:
        num = TBLBANK.objects.all().order_by('-Bkid').first().Bkid + 1
    except:
        num = 1

    for data in data_list:
        business = Business.objects.get(account__account_number=data['Bkacctno'])
        TBLBANK.objects.create(
            Bkid=num,
            Bkdivision=1,
            Mid=data['Mid'],
            Bkacctno=data['Bkacctno'],
            Bkname=data['Bkname'],
            Bkdate=data['Bkdate'][:4]+'-'+data['Bkdate'][4:6]+'-'+data['Bkdate'][6:],
            Bkjukyo=data['Bkjukyo'],
            Bkinput=data['Bkinput'],
            Bkoutput=data['Bkoutput'],
            Bkjango=data['Bkjango'],
            business=business
        )
        tr_num.value = data['Bkid']
        tr_num.save()
        num += 1

    return redirect('other_settings')

@login_required(login_url='/')
def popup_change_main_account(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    acct_list = Account.objects.filter(business=business)
    
    return render(request, 'accounting/popup_change_mainaccount.html', {'acct_list' :acct_list})

@login_required(login_url='/')
def change_main_account(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    acct_list = Account.objects.filter(business=business)
    if request.method == "POST":
        main_acctid = int(request.POST.get('acctid'))
        for acct in acct_list:
            if acct.id == main_acctid:
                acct.main = True
            else:
                acct.main = False
            acct.save()
    return HttpResponse('<script type="text/javascript">window.close(); window.opener.parent.location.reload(); window.parent.location.href="/";</script>')

@login_required(login_url='/')
def popup_transaction_division(request, Bkid):
    month = request.GET.get('month')
    tblbank_tr = TBLBANK.objects.get(Bkid=Bkid)
    total = 0
    if tblbank_tr.Bkinput:
        total=tblbank_tr.Bkinput
    elif tblbank_tr.Bkoutput:
        total=tblbank_tr.Bkoutput
    return render(request, 'accounting/popup_transaction_division.html', {'transaction': tblbank_tr, 'Bkid': Bkid, 'total': total, 'month': month})

@login_required(login_url='/')
def add_row(request):
    context = {}
    page = request.POST.get('page', None)
    if page == "popup_transaction_division":
        Bkid = request.POST.get('Bkid', None)
        tr = get_object_or_404(TBLBANK, Bkid=Bkid)
        if tr.regdatetime:
            is_regist = True
        else:
            is_regist = False
        context = {'is_regist': is_regist}
    else:
        pass
    return HttpResponse(json.dumps(context), content_type="application/json")

@login_required(login_url='/')
def delete_row(request):
    context = {}
    page = request.POST.get('page', None)
    return HttpResponse(json.dumps(context), content_type="application/json")

@login_required(login_url='/')
def regist_division(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    if request.method == "POST":
        row = int(request.POST.get('row'))
        Bkid = request.POST.get('Bkid')
        tblbank_tr = TBLBANK.objects.get(Bkid=Bkid)
        Mid = request.user.username
        inoutType = request.POST.get('inoutType')
        itemId_list = request.POST.getlist('itemId_list')
        subdivisionId_list = request.POST.getlist('subdivisionId_list')
        month = request.POST.get('month')
        Bkjukyo = request.POST.getlist('Bkjukyo')
        Bkinout = request.POST.getlist('Bkinout')
        Bkdate = request.POST.get('Bkdate')

        try:
            close = Deadline.objects.get(business=business,year=Bkdate[:4],month=Bkdate[5:7])
            if close.regdatetime:
                return HttpResponse("<script>alert('해당월은 마감완료되었습니다.');history.back();</script>")
        except:
            pass
        
        start_date = datetime.datetime.strptime(Bkdate[:8]+'01', "%Y-%m-%d")
        a_month_ago = start_date - relativedelta(months=1)
        a_month_later = start_date + relativedelta(months=1)

        #--해당날짜 전월이월금 유무확인--
        try :
            premonth_transfer = Transaction.objects.filter(business=business, Bkdate=start_date).get(Bkdivision=0)
        except IndexError:
            return HttpResponse("<script>alert('IndexError');history.back();</script>")
        except Transaction.DoesNotExist:
            #--전월이월금 없는 경우 주계좌의 이전달 마지막 내역을 전월이월금으로 등록
            main = Account.objects.filter(business=business).get(main=True)
            data = TBLBANK.objects.filter(Bkacctno=main.account_number).filter(Bkdate__gte=a_month_ago).filter(Bkdate__lt=start_date).order_by('-Bkdate', '-Bkid').first()
            if data == None:
                return HttpResponse("<script>alert('주계좌의 전월잔액이 없습니다.');history.back();</script>")
            Transaction.objects.create(
                Bkid=data.Bkid,
                Bkdivision=0,
                Mid=Mid,
                business=business,
                Bkacctno=main.account_number,
                Bkname=main.bank.name,
                Bkdate=start_date,
                Bkjukyo="전월이월금",
                Bkinput=data.Bkjango,
                Bkoutput=0,
                Bkjango=data.Bkjango,
                item=Item.objects.get(paragraph__subsection__institution=business.type3, paragraph__subsection__code=0, paragraph__code=0, code=0),
                regdatetime=today
            )

        for r in range(row, 0, -1):
            latest_tr = Transaction.objects.filter(business=business, Bkdate__lte=Bkdate).order_by('-Bkdate','-id').first()
            jango = latest_tr.Bkjango
            Bkinput = 0
            Bkoutput = 0
            Bkjango = 0
            if inoutType == "input":
                Bkinput = Bkinout[r]
                Bkjango = jango + int(Bkinout[r])
            elif inoutType == "output":
                Bkoutput = Bkinout[r]
                Bkjango = jango - int(Bkinout[r])

            item = Item.objects.get(id=itemId_list[r])

            try:
                subdivision = Subdivision.objects.get(id=subdivisionId_list[r])
            except Subdivision.DoesNotExist:
                subdivision = None
            
            Transaction.objects.create(
                Bkid=Bkid,
                Bkdivision = r,
                Mid = Mid,
                Bkacctno = tblbank_tr.Bkacctno,
                Bkname = tblbank_tr.Bkname,
                Bkdate = Bkdate,
                Bkjukyo = "[분할]"+Bkjukyo[r],
                Bkinput = Bkinput,
                Bkoutput = Bkoutput,
                Bkjango = Bkjango,
                regdatetime = today,
                item = item,
                subdivision = subdivision,
                business = business,
            )

            update_list = Transaction.objects.filter(business=business, Bkdate__gt=Bkdate, Bkdate__lt=a_month_later)
            for update in update_list:
                if inoutType == "수입":
                    update.Bkjango = int(update.Bkjango) + int(Bkinout[r])
                if inoutType == "지출":
                    update.Bkjango = int(update.Bkjango) - int(Bkinout[r])
                update.save()

            tr, created = TBLBANK.objects.get_or_create(business=business, Bkid=Bkid, Bkdivision=r, Mid=Mid, Bkacctno=tblbank_tr.Bkacctno, Bkname=tblbank_tr.Bkname, Bkdate=tblbank_tr.Bkdate)
            tr.sub_Bkjukyo="[분할]"+Bkjukyo[r]
            tr.regdatetime=today
            tr.item = item
            tr.subdivision = subdivision
            tr.save()

    return HttpResponse('<script type="text/javascript">window.close(); window.opener.parent.location.reload(); window.parent.location.href="/";</script>')

'''
@login_required(login_url='/')
def extract_subdivision(request):
    #from django.core.serializers.json import DjangoJSONEncoder
    from django.core import serializers
    item_list = request.POST.get('item_list')
    print(item_list)
    #selected_item_list = []
    """
    subdivision_list = []
    for idx, val in enumerate(item_list):
        if val:
            selected_spi_list.append(int(val))
        else:
            selected_spi_list.append(0)
    """
    subdivisions = Subdivision.objects.filter(item__id = item_list)
    html_string="<option value="">--------------</option>"
    for sub in subdivisions:
        html_string += '<option value="%s">%s</option>' % (sub.id, sub.context)
    #subdivision_list.append(subdivisions)
    #subdivision_list = json.dumps(list(subdivisions), cls=DjangoJSONEncoder)
    subdivision_list = serializers.serialize('json', list(subdivisions))
    print(html_string)
    context = {'message': '성공', 'html_string': html_string}
    return HttpResponse(json.dumps(context), content_type="application/json")
'''

@login_required(login_url='/')
def popup_select_item(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    index = request.GET.get('index')
    row = request.GET.get('row')
    inoutType = request.GET.get('inoutType')
    if inoutType == "input":
        type_filter = "수입"
    elif inoutType == "output":
        type_filter = "지출"
    from django.db import connections
    query = """SELECT s.institution_id,s.code,s.context,s.type,p.code,p.context,i.code,i.context, i.id
    FROM accounting_subsection s
    LEFT OUTER JOIN accounting_paragraph p
    ON p.subsection_id = s.id
    LEFT OUTER JOIN accounting_item i
    ON i.paragraph_id = p.id
    WHERE s.institution_id = '"""+str(business.type3)+"""' and s.type = '"""+type_filter+"""' and s.code != 0
    ORDER BY s.institution_id, s.type, s.code, p.code, i.code"""
    with connections['default'].cursor() as cursor:
        cursor.execute(query)
        spi_list = cursor.fetchall()

    subdivisions = Subdivision.objects.filter(item__paragraph__subsection__institution=business.type3)
    return render(request, 'accounting/popup_select_item.html', {'spi_list': spi_list, 'subdivisions': subdivisions, 'index': index, 'row': row})

@login_required(login_url='/')
def annual_budget(request, budget_type):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    this_year = int(DateFormat(today).format("Y"))
    if request.method == "POST":
        year = int(request.POST.get('year'))
    elif request.method == "GET":
        year = int(request.GET.get('year', this_year))

    filter_revenue = 'revenue'
    filter_expenditure = 'expenditure'
    if budget_type[:13] == "supplementary":
        filter_revenue = 'supplementary_revenue_'+budget_type[-1]
        filter_expenditure = 'supplementary_expenditure_'+budget_type[-1]

    total_revenue = Budget.objects.filter(business=business, year=year, item__paragraph__subsection__type="수입", type=filter_revenue).aggregate(total=Coalesce(Sum('price'), 0))['total']
    total_expenditure = Budget.objects.filter(business=business, year=year, item__paragraph__subsection__type="지출", type=filter_expenditure).aggregate(total=Coalesce(Sum('price'), 0))['total']
    total_difference = total_revenue - total_expenditure

    budget_list = []
    spi_list = []
    sub_budget = []
    revenue_budget_page = ''
    expenditure_budget_page = ''
    supplementary_revenue_page = ''
    supplementary_expenditure_page = ''

    if budget_type == "revenue" or budget_type[:21] == "supplementary_revenue":
        budget_list = Budget.objects.filter(business=business, year=year, item__paragraph__subsection__type="수입", type=budget_type)
        spi_list = Item.objects.filter(paragraph__subsection__institution = business.type3, paragraph__subsection__type = "수입").exclude(code=0)
        if budget_type == "revenue":
            revenue_budget_page = 'active'
        else:
            supplementary_revenue_page = 'active'
    elif budget_type == "expenditure" or budget_type[:25] == "supplementary_expenditure":
        budget_list = Budget.objects.filter(business=business, year=year, item__paragraph__subsection__type="지출", type=budget_type)
        spi_list = Item.objects.filter(paragraph__subsection__institution = business.type3, paragraph__subsection__type = "지출").exclude(code=0)
        if budget_type == "expenditure":
            expenditure_budget_page = 'active'
        else:
            supplementary_expenditure_page = 'active'

    if budget_list:
        context_list = []
        unit_price_list = []
        cnt_list = []
        months_list = []
        percent_list = []
        sub_price_list = []
        for spi in spi_list:
            spi.budget_row = 1
            for index, budget in enumerate(budget_list):
                if spi.id == budget.item.id:
                    sub_columns = ['item','context','unit_price','cnt','months','percent','sub_price']
                    context_list = budget.context.split("|")
                    unit_price_list = budget.unit_price.split("|")
                    cnt_list = budget.cnt.split("|")
                    months_list = budget.months.split("|")
                    if budget.percent is not None:
                        percent_list = budget.percent.split("|")
                    sub_price_list = budget.sub_price.split("|")
                    row_list = []
                    for idx, val in enumerate(context_list):
                        r = []
                        if val != None:
                            r.append(budget.item.id)
                            r.append(context_list[idx])
                            r.append(unit_price_list[idx])
                            r.append(cnt_list[idx])
                            r.append(months_list[idx])
                            if budget.percent is not None:
                                r.append(percent_list[idx])
                            else:
                                r.append('')
                            r.append(sub_price_list[idx])
                        row_list.append(r)
                    sub_data = [ dict(zip(sub_columns,row)) for row in row_list ]
                    spi.budget_price = budget.price
                    spi.budget_row = budget.row
                    spi.sub_budget = sub_data
                    sub_budget += sub_data

    return render(request, 'accounting/annual_budget.html', {'total_revenue': total_revenue, 'total_expenditure': total_expenditure, 'total_difference': total_difference,'budget_type': budget_type, 'sub_budget': sub_budget, 'spi_list': spi_list, 'budget_list': budget_list, 'budget_management': 'active', 'revenue_budget_page': revenue_budget_page, 'expenditure_budget_page': expenditure_budget_page, 'supplementary_revenue_page': supplementary_revenue_page, 'supplementary_expenditure_page': supplementary_expenditure_page, 'master_login': request.session['master_login'], 'business': business, 'year_range': range(this_year, 1999, -1), 'year': year})

@login_required(login_url='/')
def regist_annual_budget(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    if request.method == "POST":
        budget_type = request.POST.get('budget_type')
        budget_year = request.POST.get('budget_year')
        budget_spi = request.POST.getlist('budget_spi')
        budget_row = request.POST.getlist('budget_row')
        budget_price = request.POST.getlist('budget_price')
        budget_context = request.POST.getlist('budget_context')
        budget_unit_price = request.POST.getlist('budget_unit_price')
        budget_cnt = request.POST.getlist('budget_cnt')
        budget_months = request.POST.getlist('budget_months')
        budget_percent = request.POST.getlist('budget_percent')
        budget_sub_price = request.POST.getlist('budget_sub_price')

        row_idx = 1
        for idx, val in enumerate(budget_spi):
            if idx:
                context_list = ''
                unit_price_list = ''
                cnt_list = ''
                months_list = ''
                percent_list = ''
                sub_price_list = ''
                for row in range(1, int(budget_row[idx])+1):
                    context_list += budget_context[row_idx]+'|'
                    unit_price_list += budget_unit_price[row_idx]+'|'
                    cnt_list += budget_cnt[row_idx]+'|'
                    months_list += budget_months[row_idx]+'|'
                    percent_list += budget_percent[row_idx]+'|'
                    sub_price_list += budget_sub_price[row_idx]+'|'
                    row_idx += 1
                context_list = context_list[:-1]
                unit_price_list = unit_price_list[:-1]
                cnt_list = cnt_list[:-1]
                months_list = months_list[:-1]
                percent_list = percent_list[:-1]
                sub_price_list = sub_price_list[:-1]

                #이미 등록되어있는 경우 새로등록이 아닌 찾아 바꾸기
                budget, created = Budget.objects.get_or_create(business=business, year=budget_year, item=Item.objects.get(id=budget_spi[idx]), type=budget_type, defaults={'row': int(budget_row[idx]), 'price': 0, 'context': '', 'unit_price': '', 'cnt': '', 'months': '', 'percent': '', 'sub_price': '0'})
                budget.row = int(budget_row[idx])
                if budget_price[idx] != '': 
                    budget.price = int(budget_price[idx])
                budget.context = context_list
                budget.unit_price = unit_price_list
                budget.cnt = cnt_list
                budget.months = months_list
                budget.percent = percent_list
                budget.sub_price = sub_price_list
                budget.save()

    response = redirect('annual_budget', budget_type)
    response['Location'] += '?year='+budget_year
    return response

@login_required(login_url='/')
def budget_settlement(request, budget_type):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    this_year = int(DateFormat(today).format("Y"))
    year = this_year
    if request.method == "POST":
        year = int(request.POST.get('year'))
    if business.type3_id == "어린이집":
        start_date = datetime.datetime.strptime(str(year)+'-03-01', '%Y-%m-%d')
        end_date = datetime.datetime.strptime(str(year+1)+'-03-01', '%Y-%m-%d')
    else:
        start_date = datetime.datetime.strptime(str(year)+'-01-01', '%Y-%m-%d')
        end_date = datetime.datetime.strptime(str(year+1)+'-01-01', '%Y-%m-%d')

    revenue_settlement_page = ''
    expenditure_settlement_page = ''

    total_budget = 0
    total_sum = 0
    total_difference = 0

    if budget_type == "revenue":
        filter_type = "수입"
    elif budget_type == "expenditure":
        filter_type = "지출"

    try:
        filter_budget_type = Budget.objects.filter(business=business, year=year, type__icontains=budget_type).order_by('type').last().type
    except:
        filter_budget_type = budget_type

    subsection_list = Subsection.objects.filter(institution=business.type3, type=filter_type).annotate(count=Count('paragraph__item')).exclude(count=0).exclude(code=0)
    paragraph_list = Paragraph.objects.filter(subsection__institution=business.type3, subsection__type=filter_type).annotate(count=Count('item')).exclude(count=0)
    item_list = Item.objects.filter(paragraph__subsection__institution = business.type3, paragraph__subsection__type = filter_type).annotate(
        total_budget=Coalesce(Sum(Case(
            When(budget__business = business, then=Case(When(budget__year = year, then=Case(When(budget__type = filter_budget_type, then='budget__price'))))))), 0))
    if budget_type == "revenue":
        item_list2 = Item.objects.filter(paragraph__subsection__institution = business.type3, paragraph__subsection__type = filter_type).annotate(
            total_sum=Coalesce(Sum(Case(
                When(transaction__business = business, then=Case(
                When(transaction__Bkdate__gte = start_date, then=Case(
                When(transaction__Bkdate__lt = end_date, then='transaction__Bkinput'))))))), 0))
        revenue_settlement_page = 'active'
    elif budget_type == "expenditure":
        item_list2 = Item.objects.filter(paragraph__subsection__institution = business.type3, paragraph__subsection__type = filter_type).annotate(
            total_sum=Coalesce(Sum(Case(
                When(transaction__business = business, then=Case(
                When(transaction__Bkdate__gte = start_date, then=Case(
                When(transaction__Bkdate__lt = end_date, then='transaction__Bkoutput'))))))), 0))
        expenditure_settlement_page = 'active'

    for idx, val in enumerate(item_list):
        item_list[idx].total_sum = item_list2[idx].total_sum
        item_list[idx].total_difference = item_list[idx].total_budget - item_list[idx].total_sum
        total_budget += item_list[idx].total_budget
        total_sum += item_list[idx].total_sum
        total_difference += item_list[idx].total_difference

    return render(request,'accounting/budget_settlement.html', {'settlement_management': 'active', 'master_login': request.session['master_login'], 'revenue_settlement_page': revenue_settlement_page, 'expenditure_settlement_page': expenditure_settlement_page, 'business': business, 'year_range': range(this_year, 1999, -1), 'year': year, 'subsection_list': subsection_list, 'paragraph_list': paragraph_list, 'item_list': item_list, 'budget_type': budget_type, 'total_budget': total_budget, 'total_sum': total_sum, 'total_difference': total_difference})

@login_required(login_url='/')
def print_settlement(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    this_year = int(DateFormat(today).format("Y"))
    this_month = int(DateFormat(today).format("m"))
    year = this_year
    month = 1
    year2 = this_year
    month2 = this_month
    if business.type3_id == "어린이집":
        month = 3
        if this_month < 3:
            year = this_year - 1
    return render(request,'accounting/print_settlement.html', {'settlement_management': 'active', 'print_settlement_page': 'active', 'master_login': request.session['master_login'], 'business': business, 'year_range': range(this_year, 1999, -1), 'month_range': range(1, 13), 'year': year, 'month': month, 'year2': year2, 'month2': month2})

@login_required(login_url='/')
def print_budget_settlement(request, budget_type):
    business = get_object_or_404(Business, pk=request.session['business'])
    year = int(request.POST.get('year'))
    start_date = datetime.datetime.strptime(str(year)+'-03-01', '%Y-%m-%d')
    end_date = datetime.datetime.strptime(str(year+1)+'-03-01', '%Y-%m-%d')

    total_budget = 0
    total_sum = 0
    total_difference = 0

    if budget_type == "revenue":
        filter_type = "수입"
    elif budget_type == "expenditure":
        filter_type = "지출"

    try:
        filter_budget_type = Budget.objects.filter(business=business, year=year, type__icontains=budget_type).order_by('type').last().type
    except:
        filter_budget_type = budget_type

    subsection_list = Subsection.objects.filter(institution=business.type3, type=filter_type).annotate(count=Count('paragraph__item')).exclude(count=0).exclude(code=0)
    paragraph_list = Paragraph.objects.filter(subsection__institution=business.type3, subsection__type=filter_type).annotate(count=Count('item')).exclude(count=0)
    item_list = Item.objects.filter(paragraph__subsection__institution = business.type3, paragraph__subsection__type = filter_type).annotate(
        total_budget=Coalesce(Sum(Case(
            When(budget__business = business, then=Case(When(budget__year = year, then=Case(When(budget__type = filter_budget_type, then='budget__price'))))))), 0))
    if budget_type == "revenue":
        item_list2 = Item.objects.filter(paragraph__subsection__institution = business.type3, paragraph__subsection__type = filter_type).annotate(
            total_sum=Coalesce(Sum(Case(
                When(transaction__business = business, then=Case(
                When(transaction__Bkdate__gte = start_date, then=Case(
                When(transaction__Bkdate__lt = end_date, then='transaction__Bkinput'))))))), 0))
    elif budget_type == "expenditure":
        item_list2 = Item.objects.filter(paragraph__subsection__institution = business.type3, paragraph__subsection__type = filter_type).annotate(
            total_sum=Coalesce(Sum(Case(
                When(transaction__business = business, then=Case(
                When(transaction__Bkdate__gte = start_date, then=Case(
                When(transaction__Bkdate__lt = end_date, then='transaction__Bkoutput'))))))), 0))

    for idx, val in enumerate(item_list):
        item_list[idx].total_sum = item_list2[idx].total_sum
        item_list[idx].total_difference = item_list[idx].total_budget-item_list[idx].total_sum
        total_budget += item_list[idx].total_budget
        total_sum += item_list[idx].total_sum
        total_difference += item_list[idx].total_difference

    return render(request,'accounting/print_budget_settlement.html', {'settlement_management': 'active', 'master_login': request.session['master_login'], 'business': business, 'year': year, 'subsection_list': subsection_list, 'paragraph_list': paragraph_list, 'item_list': item_list, 'total_budget': total_budget, 'total_sum': total_sum, 'total_difference': total_difference, 'budget_type': budget_type})

@login_required(login_url='/')
def print_budget_settlement2(request, budget_type):
    business = get_object_or_404(Business, pk=request.session['business'])
    year = int(request.POST.get('year'))
    month = int(request.POST.get('month'))
    if request.POST.get('year2'):
        year2 = int(request.POST.get('year2'))
        month2 = int(request.POST.get('month2'))
    else:
        year2 = year
        month2 = month
    year3 = year2
    month3 = month2+1
    budget_year = year
    if business.type3_id == "어린이집":
        if month < 3:
            budget_year = year -1
    if month3 > 12 :
        year3 = year3 + 1
        month3 = 1
    
    if business.type3_id == "어린이집":
        session_start_date = datetime.datetime.strptime(str(budget_year)+'-03-01', '%Y-%m-%d')
    else:
        session_start_date = datetime.datetime.strptime(str(budget_year)+'-01-01', '%Y-%m-%d')

    start_date = datetime.datetime.strptime(str(year)+'-'+str(month)+'-01', '%Y-%m-%d')
    end_date = datetime.datetime.strptime(str(year3)+'-'+str(month3)+'-01', '%Y-%m-%d')

    total_budget = 0
    now_budget = 0
    total_sum = 0
    total_difference = 0

    if budget_type == "revenue":
        filter_type = "수입"
    elif budget_type == "expenditure":
        filter_type = "지출"

    try:
        filter_budget_type = Budget.objects.filter(business=business, year=year, type__icontains=budget_type).order_by('type').last().type
    except:
        filter_budget_type = budget_type

    subsection_list = Subsection.objects.filter(institution=business.type3, type=filter_type).annotate(count=Count('paragraph__item')).exclude(count=0).exclude(code=0)
    paragraph_list = Paragraph.objects.filter(subsection__institution=business.type3, subsection__type=filter_type).annotate(count=Count('item')).exclude(count=0)
    item_list = Item.objects.filter(paragraph__subsection__institution = business.type3, paragraph__subsection__type = filter_type).annotate(
        total_budget=Coalesce(Sum(Case(
            When(budget__business = business, then=Case(When(budget__year = budget_year, then=Case(When(budget__type = filter_budget_type, then='budget__price'))))))), 0))
    if budget_type == "revenue":
        item_list2 = Item.objects.filter(paragraph__subsection__institution = business.type3, paragraph__subsection__type = filter_type).annotate(
            total_sum=Coalesce(Sum(Case(
                When(transaction__business = business, then=Case(
                When(transaction__Bkdate__gte = start_date, then=Case(
                When(transaction__Bkdate__lt = end_date, then='transaction__Bkinput'))))))), 0))
        item_list3 = Item.objects.filter(paragraph__subsection__institution = business.type3, paragraph__subsection__type = filter_type).annotate(
            now_budget=Coalesce(Sum(Case(
                When(transaction__business = business, then=Case(
                When(transaction__Bkdate__gte = session_start_date, then=Case(
                When(transaction__Bkdate__lt = start_date, then='transaction__Bkinput'))))))), 0))
    elif budget_type == "expenditure":
        item_list2 = Item.objects.filter(paragraph__subsection__institution = business.type3, paragraph__subsection__type = filter_type).annotate(
            total_sum=Coalesce(Sum(Case(
                When(transaction__business = business, then=Case(
                When(transaction__Bkdate__gte = start_date, then=Case(
                When(transaction__Bkdate__lt = end_date, then='transaction__Bkoutput'))))))), 0))
        item_list3 = Item.objects.filter(paragraph__subsection__institution = business.type3, paragraph__subsection__type = filter_type).annotate(
            now_budget=Coalesce(Sum(Case(
                When(transaction__business = business, then=Case(
                When(transaction__Bkdate__gte = session_start_date, then=Case(
                When(transaction__Bkdate__lt = start_date, then='transaction__Bkoutput'))))))), 0))

    for idx, val in enumerate(item_list):
        item_list[idx].now_budget = item_list[idx].total_budget - item_list3[idx].now_budget
        item_list[idx].total_sum = item_list2[idx].total_sum
        item_list[idx].total_difference = item_list[idx].now_budget-item_list[idx].total_sum
        total_budget += item_list[idx].total_budget
        now_budget += item_list[idx].now_budget
        total_sum += item_list[idx].total_sum
        total_difference += item_list[idx].total_difference

    return render(request,'accounting/print_budget_settlement2.html', {'settlement_management': 'active', 'master_login': request.session['master_login'], 'business': business, 'year': year, 'month': month, 'year2': year2, 'month2': month2, 'subsection_list': subsection_list, 'paragraph_list': paragraph_list, 'item_list': item_list, 'total_budget': total_budget, 'total_sum': total_sum, 'total_difference': total_difference, 'budget_type': budget_type, 'now_budget': now_budget})

@login_required(login_url='/')
def print_transaction(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    year = int(request.POST.get('year'))
    month = int(request.POST.get('month'))
    if request.POST.get('year2'):
        year2 = int(request.POST.get('year2'))
        month2 = int(request.POST.get('month2'))
    else:
        year2 = year
        month2 = month
    year3 = year2
    month3 = month2+1
    if month3 > 12 :
        year3 = year3 + 1
        month3 = 1
    if business.type3_id == "어린이집":
        session_start_date = datetime.datetime.strptime(str(year)+'-03-01', '%Y-%m-%d')
    else:
        session_start_date = datetime.datetime.strptime(str(year)+'-01-01', '%Y-%m-%d')
    start_date = datetime.datetime.strptime(str(year)+'-'+str(month)+'-01', '%Y-%m-%d')
    end_date = datetime.datetime.strptime(str(year3)+'-'+str(month3)+'-01', '%Y-%m-%d')
    transaction_list = Transaction.objects.filter(business = business, Bkdate__gte = start_date, Bkdate__lt = end_date).order_by('Bkdate','Bkid','Bkdivision').exclude(item=None)

    ym_list = []
    for transaction in transaction_list:
        ym_list.append(str(transaction.Bkdate)[:7])
    ym_list = list(set(ym_list))
    ym_list.sort()

    ym_range = []
    total_input = 0
    total_output = 0
    for ym in ym_list:
        transaction = []
        ym_tr = Transaction.objects.filter(business = business, Bkdate__year=ym[:4], Bkdate__month=ym[-2:]).order_by('Bkdate','id')
        tr_paginator = Paginator(ym_tr, 40)
        for tr_page in range(1, tr_paginator.num_pages+1):
            transaction.append(tr_paginator.page(tr_page))
        
        total = Transaction.objects.filter(business=business, Bkdate__year=ym[:4], Bkdate__month=ym[-2:]).aggregate(input=Coalesce(Sum('Bkinput'),0), output=Coalesce(Sum('Bkoutput'),0))
        sub_date =  datetime.datetime.strptime(ym[:4]+'-'+ym[-2:]+'-01', '%Y-%m-%d')
        sub_end_date = DateFormat(sub_date + relativedelta(months=1)).format("Y-m-d")
        accumulated = Transaction.objects.filter(business=business, Bkdate__gte = session_start_date, Bkdate__lt=sub_end_date).aggregate(input=Coalesce(Sum('Bkinput'),0), output=Coalesce(Sum('Bkoutput'),0))

        ym_range.append({'ym': ym, 'total_input': total['input'], 'total_output': total['output'], 'transaction': transaction, 'accumulated_input': accumulated['input'], 'accumulated_output': accumulated['output']})
    
    return render(request,'accounting/print_transaction.html', {'settlement_management': 'active', 'master_login': request.session['master_login'], 'business': business, 'year': year, 'month': month, 'year2': year2, 'month2': month2, 'ym_range': ym_range, 'total_input': total_input, 'total_output':total_output})

@login_required(login_url='/')
def print_general_ledger(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    year = int(request.POST.get('year'))
    month = int(request.POST.get('month'))
    if request.POST.get('year2'):
        year2 = int(request.POST.get('year2'))
        month2 = int(request.POST.get('month2'))
    else:
        year2 = year
        month2 = month
    year3 = year2
    month3 = month2+1
    budget_year = year
    if month3 > 12 :
        year3 = year3 + 1
        month3 = 1
    session_start_date = datetime.datetime.strptime(str(year)+'-03-01', '%Y-%m-%d')
    start_date = datetime.datetime.strptime(str(year)+'-'+str(month)+'-01', '%Y-%m-%d')
    end_date = datetime.datetime.strptime(str(year3)+'-'+str(month3)+'-01', '%Y-%m-%d')

    if business.type3_id == "어린이집" and month < 3:
        budget_year = year - 1

    item_list = Item.objects.filter(paragraph__subsection__institution = business.type3).exclude(paragraph__subsection__code=0)
    for item in item_list :
        transaction = []
        tr_in_item = Transaction.objects.filter(business = business, Bkdate__gte = start_date, Bkdate__lt = end_date, item = item).order_by('Bkdate','Bkid','Bkdivision')
        tr_total = Transaction.objects.filter(business = business, Bkdate__gte = start_date, Bkdate__lt = end_date, item = item).order_by('Bkdate','Bkid','Bkdivision').aggregate(Bkinput=Coalesce(Sum('Bkinput'),0), Bkoutput=Coalesce(Sum('Bkoutput'),0))
        tr_paginator = Paginator(tr_in_item, 40)
        for tr_page in range(1, tr_paginator.num_pages+1):
            transaction.append(tr_paginator.page(tr_page))
        if item.paragraph.subsection.type == "수입":
            filter_type = "revenue"
        elif item.paragraph.subsection.type == "지출":
            filter_type = "expenditure"
        budget = Budget.objects.filter(business = business, year = budget_year, item=item, type__icontains=filter_type).order_by('type').last()
        item.transaction = transaction
        item.total_input = tr_total['Bkinput']
        item.total_output = tr_total['Bkoutput']
        if budget != None:
            item.budget = budget.price
        else:
            item.budget = 0
        item.balance = item.budget - item.total_input - item.total_output

    return render(request,'accounting/print_general_ledger.html', {'settlement_management': 'active', 'master_login': request.session['master_login'], 'business': business, 'year': year, 'month': month, 'year2': year2, 'month2': month2, 'item_list': item_list})

@login_required(login_url='/')
def print_voucher(request, voucher_type):
    business = get_object_or_404(Business, pk=request.session['business'])
    year = int(request.POST.get('year'))
    month = int(request.POST.get('month'))
    if request.POST.get('year2'):
        year2 = int(request.POST.get('year2'))
        month2 = int(request.POST.get('month2'))
    else:
        year2 = year
        month2 = month
    year3 = year2
    month3 = month2+1
    if month3 > 12 :
        year3 = year3 + 1
        month3 = 1
    session_start_date = datetime.datetime.strptime(str(year)+'-03-01', '%Y-%m-%d')
    start_date = datetime.datetime.strptime(str(year)+'-'+str(month)+'-01', '%Y-%m-%d')
    end_date = datetime.datetime.strptime(str(year3)+'-'+str(month3)+'-01', '%Y-%m-%d')

    if voucher_type == 'revenue':
        filter_type = "수입"
        transaction_list = Transaction.objects.filter(business=business, Bkdate__gte=start_date, Bkdate__lt=end_date, Bkinput__gt=0)
    else:
        filter_type = "지출"
        transaction_list = Transaction.objects.filter(business=business, Bkdate__gte=start_date, Bkdate__lt=end_date, Bkoutput__gt=0)

    ymd_list = []
    for transaction in transaction_list:
        ymd_list.append(transaction.Bkdate)
    ymd_list = list(set(ymd_list))
    ymd_list.sort()
    

    data_list = []
    for ymd in ymd_list:
        item_list = Item.objects.filter(transaction__business=business, transaction__Bkdate=ymd, paragraph__subsection__type=filter_type).exclude(paragraph__subsection__code=0).distinct()
        for item in item_list:
            transaction = []
            if voucher_type == 'revenue':
                tr_in_item = Transaction.objects.filter(business=business, Bkdate=ymd, item=item, Bkinput__gt=0)
            else:
                tr_in_item = Transaction.objects.filter(business=business, Bkdate=ymd, item=item, Bkoutput__gt=0)
            tr_sum = tr_in_item.aggregate(input=Coalesce(Sum('Bkinput'),0), output=Coalesce(Sum('Bkoutput'),0))
            tr_paginator = Paginator(tr_in_item, 20)
            for tr_page in range(1, tr_paginator.num_pages+1):
                transaction.append(tr_paginator.page(tr_page))

            item.transaction = transaction
            item.sum = tr_sum['input'] + tr_sum['output']
            item.sum_ko = readNumber(str(item.sum))
            data_list.append({'date': ymd, 'item': item})

    return render(request,'accounting/print_voucher.html', {'settlement_management': 'active', 'master_login': request.session['master_login'], 'business': business, 'year': year, 'month': month, 'year2': year2, 'month2': month2, 'data_list': data_list, 'voucher_type': voucher_type})

@login_required(login_url='/')
def popup_returned_transaction(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    month = request.GET.get('month')
    Bkid = request.GET.get('Bkid')
    acctid = request.GET.get('acctid')
    Bkdate = request.GET.get('Bkdate')
    acct = get_object_or_404(Account, business=business, id=acctid)
    tr = TBLBANK.objects.get(Mid=business.owner.profile.user.username, Bkid=Bkid)
    if tr.Bkinput > 0:
        opposite_type = "지출"
    else :
        opposite_type = "수입"
    item_list = Item.objects.filter(paragraph__subsection__type=opposite_type, paragraph__subsection__institution=business.type3).exclude(paragraph__subsection__code=0)

    return render(request,'accounting/popup_returned_transaction.html', {'acct': acct, 'transaction': tr, 'item_list': item_list, 'month': month, 'Bkdate': Bkdate})

@login_required(login_url='/')
def regist_returned_transaction(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    month = request.POST.get('month')
    Bkid = request.POST.get('Bkid')
    acctid = request.POST.get('acctid')
    Bkdate = request.POST.get('Bkdate')
    remark = request.POST.get('remark')
    item = Item.objects.get(id=request.POST.get('item'))
    tr = TBLBANK.objects.get(Mid=business.owner.profile.user.username, Bkid=Bkid)
    acct = get_object_or_404(Account, business=business, id=acctid)

    try:
        close = Deadline.objects.get(business=business,year=Bkdate[:4],month=Bkdate[5:7])
        if close.regdatetime:
            return HttpResponse("<script>alert('해당월은 마감완료되었습니다.');history.back();</script>")
    except:
        pass

    #--해당날짜 전월이월금 유무확인--
    start_date = datetime.datetime.strptime(Bkdate[:8]+'01', "%Y-%m-%d")
    a_month_ago = start_date - relativedelta(months=1)
    a_month_later = start_date + relativedelta(months=1)

    try :
        carryover_tr = Transaction.objects.filter(business=business, Bkdate=start_date).get(Bkdivision=0)
    except IndexError:
        return HttpResponse("<script>alert('IndexError');history.back();</script>")
    except Transaction.DoesNotExist:
        #--전월이월금 없는 경우 주계좌의 이전달 마지막 내역을 전월이월금으로 등록
        main_acct = Account.objects.filter(business=business).get(main=True)
        last_tr = TBLBANK.objects.filter(Bkacctno=main_acct.account_number, Bkdate__gte=a_month_ago).filter(Bkdate__lt=start_date).order_by('Bkdate','Bkid').last()
        if last_tr == None:
            return HttpResponse("<script>alert('주계좌의 전월 거래내역이 없습니다. 전월거래내역이 있는 계좌를 주계좌로 변경하세요.');history.back();</script>")
        Transaction.objects.create(
            Bkid=last_tr.Bkid,
            Bkdivision=0,
            Mid=business.owner.profile.user.username,
            business=business,
            Bkacctno=main_acct.account_number,
            Bkname=main_acct.bank.name,
            Bkdate=start_date,
            Bkjukyo="전월이월금",
            Bkinput=last_tr.Bkjango,
            Bkoutput=0,
            Bkjango=last_tr.Bkjango,
            item=Item.objects.get(paragraph__subsection__institution=business.type3, paragraph__subsection__code=0, paragraph__code=0, code=0),
            regdatetime=today
        )
    
    latest_tr = Transaction.objects.filter(business=business, Bkdate__lte=Bkdate).order_by('-Bkdate','-id').first()
    if tr.Bkinput > 0 :
        Bkinput = 0
        Bkoutput = tr.Bkinput * -1
        Bkjango = latest_tr.Bkjango - Bkoutput
    else :
        Bkinput = tr.Bkoutput * -1
        Bkoutput = 0
        Bkjango = latest_tr.Bkjango + Bkinput
    
    transaction = Transaction(
        Bkid=tr.Bkid,
        Bkdivision=1,
        Mid=business.owner.profile.user.username,
        business=business,
        Bkacctno=acct.account_number,
        Bkname=acct.bank.name,
        Bkdate=Bkdate,
        Bkjukyo=request.POST.get('Bkjukyo'),
        Bkinput=Bkinput,
        Bkoutput=Bkoutput,
        Bkjango=Bkjango,
        regdatetime=today,
        item = item,
        remark = remark+"[반납결의서]"
    )
    transaction.save()

    update_list = Transaction.objects.filter(business=business, Bkdate__gt=Bkdate, Bkdate__lt=a_month_later)
    for update in update_list:
        if transaction.Bkinput < 0:
            update.Bkjango = update.Bkjango + transaction.Bkinput
        elif transaction.Bkoutput < 0:
            update.Bkjango = update.Bkjango - transaction.Bkoutput
        update.save()

    tr_update = TBLBANK.objects.get(Bkid=Bkid)
    tr_update.sub_Bkjukyo=request.POST.get('Bkjukyo')+' - '+remark+'[반납결의서]'
    tr_update.item = item
    tr_update.regdatetime=today
    tr_update.save()

    return HttpResponse('<script type="text/javascript">window.close(); window.opener.parent.location.reload(); window.parent.location.href="/";</script>')

@login_required(login_url='/')
def popup_transaction_direct(request):
    today = datetime.datetime.now()
    year = request.GET.get('year')
    month = request.GET.get('month')
    Bkdate = datetime.datetime.strptime(year+'-'+month+'-01', "%Y-%m-%d")
    business = get_object_or_404(Business, pk=request.session['business'])
    tblbankform = TblbankDirectForm(initial={'Bkdate':Bkdate})
    tblbankform.fields['item'].queryset = Item.objects.filter(paragraph__subsection__institution = business.type3, paragraph__subsection__type="수입")
    return render(request,'accounting/popup_transaction_direct.html', {'transactionform': tblbankform, 'year': year, 'month': month})

@login_required(login_url='/')
def change_item_option(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    inoutType = request.GET.get('inoutType')
    year = request.GET.get('year')
    month = request.GET.get('month')
    Bkdate = datetime.datetime.strptime(year+'-'+month+'-01', "%Y-%m-%d")
    html = '<option value="" selected="">---------</option>'
    if inoutType == "input":
        type_filter = "수입"
    elif inoutType == "output":
        type_filter = "지출"
    try :
        carryover_tr = Transaction.objects.filter(business=business, Bkdate=Bkdate).get(Bkdivision=0)
        items = Item.objects.filter(paragraph__subsection__institution = business.type3, paragraph__subsection__type=type_filter).exclude(paragraph__subsection__code=0)
    except Transaction.DoesNotExist:
        items = Item.objects.filter(paragraph__subsection__institution = business.type3, paragraph__subsection__type=type_filter)
    for item in items:
        html += '<option value="'+str(item.id)+'">'+str(item)+'</option>'
    context = {'item': html}
    return HttpResponse(json.dumps(context), content_type="application/json")

@login_required(login_url='/')
def regist_transaction_direct(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    inout_type = request.POST.get('inout')
    remark = request.POST.get('remark')
    tblbankform = TblbankDirectForm(request.POST)
    Bkdate = request.POST.get('Bkdate')

    try:
        close = Deadline.objects.get(business=business,year=Bkdate[:4],month=Bkdate[5:7])
        if close.regdatetime:
            return HttpResponse("<script>alert('해당월은 마감완료되었습니다.');history.back();</script>")
    except:
        pass
    
    Bkid = TBLBANK.objects.all().order_by('-Bkid').first().Bkid + 1
    if tblbankform.is_valid():
        tr = tblbankform.save(commit=False)
        tr.business = business
        tr.Bkid = Bkid
        if tr.item.context == "전월이월금":
            Bkdivision = 0
        else:
            Bkdivision = 1
        tr.Bkdivision = Bkdivision
        tr.Mid = business.owner.profile.user.username
        tr.regdatetime = today
        tr.direct = True
        tr.save()

        Bkdate =  tr.Bkdate
        start_date = datetime.datetime.strptime(DateFormat(tr.Bkdate).format("Y-m-01"), '%Y-%m-%d')
        #print(start_date) #2018-09-01 00:00:00
        #start_date2 = tr.Bkdate
        #print(start_date2) #2019-02-13 00:00:00+09:00
        a_month_ago = start_date - relativedelta(months=1)
        a_month_later = start_date + relativedelta(months=1)

        #--전월 이월금 등록--
        try :
            carryover_tr = Transaction.objects.filter(business=business, Bkdate=start_date).get(Bkdivision=0)
        #--전월이월금 없는 경우
        except Transaction.DoesNotExist:
            #--등록되는 거래가 전월이월금이면 transaction테이블에 등록 후 창닫기
            if tr.item.context == "전월이월금":
                Transaction.objects.create(
                    Bkid=Bkid,                  Bkdivision=0,
                    Mid=business.owner.profile.user.username,  business=business,
                    Bkdate=start_date,          Bkjukyo="전월이월금",
                    Bkinput=tr.Bkinput,         Bkoutput=0,
                    Bkjango=tr.Bkinput,         regdatetime=today,
                    item=Item.objects.get(paragraph__subsection__institution=business.type3, paragraph__subsection__code=0, paragraph__code=0, code=0),
                    remark=remark
                )
                return HttpResponse('<script type="text/javascript">window.close(); window.opener.parent.location.reload(); window.parent.location.href="/";</script>')
            #--주계좌의 이전달 마지막 내역을 전월이월금으로 등록
            else:
                main_acct = Account.objects.filter(business=business).get(main=True)
                last_tr = TBLBANK.objects.filter(Bkacctno=main_acct.account_number, Bkdate__gte=a_month_ago).filter(Bkdate__lt=start_date).order_by('Bkdate','Bkid').last()
                #--전월 주계좌 거래내역 없으면 메시지 출력
                if last_tr == None:
                    tr.delete()
                    return HttpResponse("<script>alert('주계좌의 전월 거래내역이 없습니다. 주계좌를 전월거래내역이 있는 계좌로 변경하거나 전월이월금을 직접등록한 후 이용해주세요.');history.back();</script>")
                Transaction.objects.create(
                    Bkid=last_tr.Bkid,          Bkdivision=0,
                    Mid=business.owner.profile.user.username,  business=business,
                    Bkacctno=main_acct.account_number,
                    Bkname=main_acct.bank.name, Bkdate=start_date,
                    Bkjukyo="전월이월금",       Bkinput=last_tr.Bkjango,
                    Bkoutput=0,                 Bkjango=last_tr.Bkjango,
                    item=Item.objects.get(paragraph__subsection__institution=business.type3, paragraph__subsection__code=0, paragraph__code=0, code=0),
                    regdatetime=today
                )
        #--전월이월금 등록 완료

        #--거래등록
        if tr.item.context == "전월이월금":
            return HttpResponse("<script>alert('해당월의 전년도 이월금은 이미 등록되었습니다.');history.back();</script>")

        latest_tr = Transaction.objects.filter(business=business, Bkdate__lte=Bkdate).order_by('-Bkdate','-id').first()
        if inout_type == "input":
            Bkjango = latest_tr.Bkjango + tr.Bkinput
        elif inout_type == "output":
            Bkjango = latest_tr.Bkjango - tr.Bkoutput

        transaction = Transaction(
            Bkid=Bkid,                  Bkdivision=Bkdivision,
            Mid=business.owner.profile.user.username,    business=business,
            Bkdate=Bkdate,              Bkjukyo=tr.Bkjukyo,
            Bkinput=tr.Bkinput,         Bkoutput=tr.Bkoutput,
            Bkjango=Bkjango,            regdatetime=today,
            item = tr.item,             remark = remark
        )
        transaction.save()

        update_list = Transaction.objects.filter(business=business, Bkdate__gt=Bkdate, Bkdate__lt=a_month_later)
        for update in update_list:
            if transaction.Bkinput > 0:
                update.Bkjango = update.Bkjango + transaction.Bkinput
            elif transaction.Bkoutput > 0:
                update.Bkjango = update.Bkjango - transaction.Bkoutput
            update.save()

    return HttpResponse('<script type="text/javascript">window.close(); window.opener.parent.location.reload(); window.parent.location.href="/";</script>')

@login_required(login_url='/')
def print_returned_voucher(request, voucher_type):
    business = get_object_or_404(Business, pk=request.session['business'])
    year = int(request.POST.get('year'))
    month = int(request.POST.get('month'))
    if request.POST.get('year2'):
        year2 = int(request.POST.get('year2'))
        month2 = int(request.POST.get('month2'))
    else:
        year2 = year
        month2 = month
    year3 = year2
    month3 = month2+1
    if month3 > 12 :
        year3 = year3 + 1
        month3 = 1
    session_start_date = datetime.datetime.strptime(str(year)+'-03-01', '%Y-%m-%d')
    start_date = datetime.datetime.strptime(str(year)+'-'+str(month)+'-01', '%Y-%m-%d')
    end_date = datetime.datetime.strptime(str(year3)+'-'+str(month3)+'-01', '%Y-%m-%d')

    if voucher_type == 'revenue':
        filter_type = "수입"
        transaction_list = Transaction.objects.filter(business=business, Bkdate__gte=start_date, Bkdate__lt=end_date, Bkinput__lt=0)
    else:
        filter_type = "지출"
        transaction_list = Transaction.objects.filter(business=business, Bkdate__gte=start_date, Bkdate__lt=end_date, Bkoutput__lt=0)

    ymd_list = []
    for transaction in transaction_list:
        ymd_list.append(transaction.Bkdate)
    ymd_list = list(set(ymd_list))
    ymd_list.sort()
    

    data_list = []
    for ymd in ymd_list:
        item_list = Item.objects.filter(transaction__business=business, transaction__Bkdate=ymd, paragraph__subsection__type=filter_type).exclude(paragraph__subsection__code=0).distinct()
        for item in item_list:
            if voucher_type == 'revenue':
                transaction = Transaction.objects.filter(business=business, Bkdate=ymd, item=item, Bkinput__lt=0)
            else:
                transaction = Transaction.objects.filter(business=business, Bkdate=ymd, item=item, Bkoutput__lt=0)
            for tr_list in transaction:
                tr_list.Bkinput = tr_list.Bkinput * -1
                tr_list.Bkoutput = tr_list.Bkoutput * -1
            tr_sum = transaction.aggregate(input=Coalesce(Sum('Bkinput'),0), output=Coalesce(Sum('Bkoutput'),0))
            item.transaction = transaction
            item.sum = tr_sum['input'] * -1 + tr_sum['output'] * -1
            item.sum_ko = readNumber(str(item.sum))
            data_list.append({'date': ymd, 'item': item})

    return render(request,'accounting/print_returned_voucher.html', {'settlement_management': 'active', 'master_login': request.session['master_login'], 'business': business, 'year': year, 'month': month, 'year2': year2, 'month2': month2, 'data_list': data_list, 'voucher_type': voucher_type})

@login_required(login_url='/')
def monthly_print(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    this_year = int(DateFormat(today).format("Y"))
    this_month = int(DateFormat(today).format("m"))
    year = this_year
    month = 3
    year2 = this_year
    month2 = this_month
    if this_month < 3:
        year = this_year - 1
    return render(request,'accounting/monthly_print.html', {'accounting_management': 'active', 'monthly_print': 'active', 'master_login': request.session['master_login'], 'business': business, 'year_range': range(this_year, 1999, -1), 'month_range': range(1, 13), 'year': year, 'month': month, 'year2': year2, 'month2': month2})

@login_required(login_url='/')
def monthly_print_all(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    #--------기간설정----------
    year = int(request.POST.get('year'))
    month = int(request.POST.get('month'))
    if request.POST.get('year2'):
        year2 = int(request.POST.get('year2'))
        month2 = int(request.POST.get('month2'))
    else:
        year2 = year
        month2 = month
    year3 = year2
    month3 = month2+1
    budget_year = year
    if business.type3_id == "어린이집":
        if month < 3:
            budget_year = year -1
    if month3 > 12 :
        year3 = year3 + 1
        month3 = 1
    start_date = datetime.datetime.strptime(str(year)+'-'+str(month)+'-01', '%Y-%m-%d')
    end_date = datetime.datetime.strptime(str(year3)+'-'+str(month3)+'-01', '%Y-%m-%d')

    #--------세입결산----------
    revenue_total_budget = 0
    revenue_total_sum = 0
    revenue_total_difference = 0

    try:
        filter_budget_type = Budget.objects.filter(business=business, year=year, type__icontains="revenue").order_by('type').last().type
    except:
        filter_budget_type = "revenue"

    revenue_subsection_list = Subsection.objects.filter(institution=business.type3, type="수입").annotate(count=Count('paragraph__item')).exclude(count=0).exclude(code=0)
    revenue_paragraph_list = Paragraph.objects.filter(subsection__institution=business.type3, subsection__type="수입").annotate(count=Count('item')).exclude(count=0)
    revenue_item_list = Item.objects.filter(paragraph__subsection__institution = business.type3, paragraph__subsection__type = "수입").annotate(
        total_budget=Coalesce(Sum(Case(
            When(budget__business = business, then=Case(When(budget__year = budget_year, then=Case(When(budget__type = filter_budget_type, then='budget__price'))))))), 0))
    revenue_item_list2 = Item.objects.filter(paragraph__subsection__institution = business.type3, paragraph__subsection__type = "수입").annotate(
            total_sum=Coalesce(Sum(Case(
                When(transaction__business = business, then=Case(
                When(transaction__Bkdate__gte = start_date, then=Case(
                When(transaction__Bkdate__lt = end_date, then='transaction__Bkinput'))))))), 0))

    for idx, val in enumerate(revenue_item_list):
        revenue_item_list[idx].total_sum = revenue_item_list2[idx].total_sum
        revenue_item_list[idx].total_difference = revenue_item_list[idx].total_budget-revenue_item_list[idx].total_sum
        revenue_total_budget += revenue_item_list[idx].total_budget
        revenue_total_sum += revenue_item_list[idx].total_sum
        revenue_total_difference += revenue_item_list[idx].total_difference

    #--------세출결산----------
    expenditure_total_budget = 0
    expenditure_total_sum = 0
    expenditure_total_difference = 0

    try:
        filter_budget_type = Budget.objects.filter(business=business, year=year, type__icontains="expenditure").order_by('type').last().type
    except:
        filter_budget_type = "expenditure"

    expenditure_subsection_list = Subsection.objects.filter(institution=business.type3, type="지출").annotate(count=Count('paragraph__item')).exclude(count=0)
    expenditure_paragraph_list = Paragraph.objects.filter(subsection__institution=business.type3, subsection__type="지출").annotate(count=Count('item')).exclude(count=0)
    expenditure_item_list = Item.objects.filter(paragraph__subsection__institution = business.type3, paragraph__subsection__type = "지출").annotate(
        total_budget=Coalesce(Sum(Case(
            When(budget__business = business, then=Case(When(budget__year = budget_year, then=Case(When(budget__type = filter_budget_type, then='budget__price'))))))), 0))
    expenditure_item_list2 = Item.objects.filter(paragraph__subsection__institution = business.type3, paragraph__subsection__type = "지출").annotate(
            total_sum=Coalesce(Sum(Case(
                When(transaction__business = business, then=Case(
                When(transaction__Bkdate__gte = start_date, then=Case(
                When(transaction__Bkdate__lt = end_date, then='transaction__Bkoutput'))))))), 0))

    for idx, val in enumerate(expenditure_item_list):
        expenditure_item_list[idx].total_sum = expenditure_item_list2[idx].total_sum
        expenditure_item_list[idx].total_difference = expenditure_item_list[idx].total_budget-expenditure_item_list[idx].total_sum
        expenditure_total_budget += expenditure_item_list[idx].total_budget
        expenditure_total_sum += expenditure_item_list[idx].total_sum
        expenditure_total_difference += expenditure_item_list[idx].total_difference

    #--------현금출납부----------
    if business.type3_id == "어린이집":
        session_start_date = datetime.datetime.strptime(str(year)+'-03-01', '%Y-%m-%d')
    else:
        session_start_date = datetime.datetime.strptime(str(year)+'-01-01', '%Y-%m-%d')

    transaction_list = Transaction.objects.filter(business = business, Bkdate__gte = start_date, Bkdate__lt = end_date).order_by('Bkdate','Bkid','Bkdivision').exclude(item=None)

    ym_list = []
    for transaction in transaction_list:
        ym_list.append(str(transaction.Bkdate)[:7])
    ym_list = list(set(ym_list))
    ym_list.sort()

    ym_range = []
    total_input = 0
    total_output = 0
    for ym in ym_list:
        transaction = []
        ym_tr = Transaction.objects.filter(business = business, Bkdate__year=ym[:4], Bkdate__month=ym[-2:]).order_by('Bkdate','Bkid')
        tr_paginator = Paginator(ym_tr, 40)
        for tr_page in range(1, tr_paginator.num_pages+1):
            transaction.append(tr_paginator.page(tr_page))
        
        total = Transaction.objects.filter(business=business, Bkdate__year=ym[:4], Bkdate__month=ym[-2:]).aggregate(input=Coalesce(Sum('Bkinput'),0), output=Coalesce(Sum('Bkoutput'),0))
        sub_date =  datetime.datetime.strptime(ym[:4]+'-'+ym[-2:]+'-01', '%Y-%m-%d')
        sub_end_date = DateFormat(sub_date + relativedelta(months=1)).format("Y-m-d")
        accumulated = Transaction.objects.filter(business=business, Bkdate__gte = session_start_date, Bkdate__lt=sub_end_date).aggregate(input=Coalesce(Sum('Bkinput'),0), output=Coalesce(Sum('Bkoutput'),0))

        ym_range.append({'ym': ym, 'total_input': total['input'], 'total_output': total['output'], 'transaction': transaction, 'accumulated_input': accumulated['input'], 'accumulated_output': accumulated['output']})

    #--------총계정원장----------
    general_ledger_list = Item.objects.filter(paragraph__subsection__institution = business.type3).exclude(paragraph__subsection__code=0)
    for general_ledger in general_ledger_list :
        transaction = []
        tr_in_general_ledger = Transaction.objects.filter(business = business, Bkdate__gte = start_date, Bkdate__lt = end_date, item = general_ledger).order_by('Bkdate','Bkid','Bkdivision')
        tr_total = Transaction.objects.filter(business = business, Bkdate__gte = start_date, Bkdate__lt = end_date, item = general_ledger).order_by('Bkdate','Bkid','Bkdivision').aggregate(Bkinput=Coalesce(Sum('Bkinput'),0), Bkoutput=Coalesce(Sum('Bkoutput'),0))
        tr_paginator = Paginator(tr_in_general_ledger, 40)
        for tr_page in range(1, tr_paginator.num_pages+1):
            transaction.append(tr_paginator.page(tr_page))
        if general_ledger.paragraph.subsection.type == "수입":
            filter_type = "revenue"
        elif general_ledger.paragraph.subsection.type == "지출":
            filter_type = "expenditure"
        budget = Budget.objects.filter(business = business, year = budget_year, item=general_ledger, type__icontains=filter_type).order_by('type').last()
        general_ledger.transaction = transaction
        general_ledger.total_input = tr_total['Bkinput']
        general_ledger.total_output = tr_total['Bkoutput']
        if budget != None:
            general_ledger.budget = budget.price
        else:
            general_ledger.budget = 0
        general_ledger.balance = general_ledger.budget - general_ledger.total_input - general_ledger.total_output

    #--------수입결의서----------
    transaction_list = Transaction.objects.filter(business=business, Bkdate__gte=start_date, Bkdate__lt=end_date, Bkinput__gt=0)

    ymd_list = []
    for transaction in transaction_list:
        ymd_list.append(transaction.Bkdate)
    ymd_list = list(set(ymd_list))
    ymd_list.sort()

    revenue_voucher_list = []
    for ymd in ymd_list:
        item_list = Item.objects.filter(transaction__business=business, transaction__Bkdate=ymd, paragraph__subsection__type="수입").exclude(paragraph__subsection__code=0).distinct()
        for item in item_list:
            transaction = []
            tr_in_item = Transaction.objects.filter(business=business, Bkdate=ymd, item=item, Bkinput__gt=0)
            tr_paginator = Paginator(tr_in_item, 20)
            for tr_page in range(1, tr_paginator.num_pages+1):
                transaction.append(tr_paginator.page(tr_page))
            tr_sum = tr_in_item.aggregate(input=Coalesce(Sum('Bkinput'),0), output=Coalesce(Sum('Bkoutput'),0))
            item.transaction = transaction
            item.sum = tr_sum['input'] + tr_sum['output']
            item.sum_ko = readNumber(str(item.sum))
            revenue_voucher_list.append({'date': ymd, 'item': item})

    #--------지출결의서----------
    transaction_list = Transaction.objects.filter(business=business, Bkdate__gte=start_date, Bkdate__lt=end_date, Bkoutput__gt=0)

    ymd_list = []
    for transaction in transaction_list:
        ymd_list.append(transaction.Bkdate)
    ymd_list = list(set(ymd_list))
    ymd_list.sort()

    expenditure_voucher_list = []
    for ymd in ymd_list:
        item_list = Item.objects.filter(transaction__business=business, transaction__Bkdate=ymd, paragraph__subsection__type="지출").exclude(paragraph__subsection__code=0).distinct()
        for item in item_list:
            transaction = []
            tr_in_item = Transaction.objects.filter(business=business, Bkdate=ymd, item=item, Bkoutput__gt=0)
            tr_paginator = Paginator(tr_in_item, 20)
            for tr_page in range(1, tr_paginator.num_pages+1):
                transaction.append(tr_paginator.page(tr_page))
            tr_sum = tr_in_item.aggregate(input=Coalesce(Sum('Bkinput'),0), output=Coalesce(Sum('Bkoutput'),0))
            item.transaction = transaction
            item.sum = tr_sum['input'] + tr_sum['output']
            item.sum_ko = readNumber(str(item.sum))
            expenditure_voucher_list.append({'date': ymd, 'item': item})

    #--------수입반납결의서----------
    transaction_list = Transaction.objects.filter(business=business, Bkdate__gte=start_date, Bkdate__lt=end_date, Bkinput__lt=0)

    ymd_list = []
    for transaction in transaction_list:
        ymd_list.append(transaction.Bkdate)
    ymd_list = list(set(ymd_list))
    ymd_list.sort()

    revenue_returned_voucher_list = []
    for ymd in ymd_list:
        item_list = Item.objects.filter(transaction__business=business, transaction__Bkdate=ymd, paragraph__subsection__type="수입").exclude(paragraph__subsection__code=0).distinct()
        for item in item_list:
            transaction = Transaction.objects.filter(business=business, Bkdate=ymd, item=item, Bkinput__lt=0)
            for tr_list in transaction:
                tr_list.Bkinput = tr_list.Bkinput * -1
                tr_list.Bkoutput = tr_list.Bkoutput * -1
            tr_sum = transaction.aggregate(input=Coalesce(Sum('Bkinput'),0), output=Coalesce(Sum('Bkoutput'),0))
            item.transaction = transaction
            item.sum = tr_sum['input'] * -1 + tr_sum['output'] * -1
            item.sum_ko = readNumber(str(item.sum))
            revenue_returned_voucher_list.append({'date': ymd, 'item': item})

    #--------지출반납결의서----------
    transaction_list = Transaction.objects.filter(business=business, Bkdate__gte=start_date, Bkdate__lt=end_date, Bkoutput__lt=0)

    ymd_list = []
    for transaction in transaction_list:
        ymd_list.append(transaction.Bkdate)
    ymd_list = list(set(ymd_list))
    ymd_list.sort()

    expenditure_returned_voucher_list = []
    for ymd in ymd_list:
        item_list = Item.objects.filter(transaction__business=business, transaction__Bkdate=ymd, paragraph__subsection__type="지출").exclude(paragraph__subsection__code=0).distinct()
        for item in item_list:
            transaction = Transaction.objects.filter(business=business, Bkdate=ymd, item=item, Bkoutput__lt=0)
            for tr_list in transaction:
                tr_list.Bkinput = tr_list.Bkinput * -1
                tr_list.Bkoutput = tr_list.Bkoutput * -1
            tr_sum = transaction.aggregate(input=Coalesce(Sum('Bkinput'),0), output=Coalesce(Sum('Bkoutput'),0))
            item.transaction = transaction
            item.sum = tr_sum['input'] * -1 + tr_sum['output'] * -1
            item.sum_ko = readNumber(str(item.sum))
            expenditure_returned_voucher_list.append({'date': ymd, 'item': item})

    return render(request,'accounting/monthly_print_all.html', {
        'settlement_management': 'active', 'master_login': request.session['master_login'],
        'business': business, 'year': year, 'month': month, 'year2': year2, 'month2': month2,
        'revenue_subsection_list': revenue_subsection_list, 'revenue_paragraph_list': revenue_paragraph_list, 'revenue_item_list': revenue_item_list, 'revenue_total_budget': revenue_total_budget, 'revenue_total_sum': revenue_total_sum, 'revenue_total_difference': revenue_total_difference,
        'expenditure_subsection_list': expenditure_subsection_list, 'expenditure_paragraph_list': expenditure_paragraph_list, 'expenditure_item_list': expenditure_item_list, 'expenditure_total_budget': expenditure_total_budget, 'expenditure_total_sum': expenditure_total_sum, 'expenditure_total_difference': expenditure_total_difference,
        'ym_range': ym_range, 'total_input': total_input, 'total_output':total_output,
        'general_ledger_list': general_ledger_list,
        'revenue_voucher_list': revenue_voucher_list,
        'expenditure_voucher_list': expenditure_voucher_list,
        'revenue_returned_voucher_list': revenue_returned_voucher_list,
        'expenditure_returned_voucher_list': expenditure_returned_voucher_list
    })

@login_required(login_url='/')
def close_list(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    if request.method == "GET":
        today = datetime.datetime.now()
        this_year = today.year
        selected_year = int(request.GET.get('year', this_year))

        if business.type3_id == "어린이집":
            session_start_date = datetime.datetime.strptime(str(selected_year)+'-03-01', '%Y-%m-%d')
            session_end_date = datetime.datetime.strptime(str(selected_year+1)+'-03-01', '%Y-%m-%d')
            row_list = [ [selected_year+1, 1], [selected_year+1, 2] ]
        else:
            session_start_date = datetime.datetime.strptime(str(selected_year)+'-01-01', '%Y-%m-%d')
            session_end_date = datetime.datetime.strptime(str(selected_year+1)+'-01-01', '%Y-%m-%d')
            row_list = [ [selected_year, 1], [selected_year, 2] ]

    #ym_list = TBLBANK.objects.filter(Bkdate__gte=session_start_date, Bkdate__lt=session_end_date).values('Bkdate__year', 'Bkdate__month').order_by('Bkdate__year','Bkdate__month').distinct()
        row_list.append([selected_year, 3])
        row_list.append([selected_year, 4])
        row_list.append([selected_year, 5])
        row_list.append([selected_year, 6])
        row_list.append([selected_year, 7])
        row_list.append([selected_year, 8])
        row_list.append([selected_year, 9])
        row_list.append([selected_year, 10])
        row_list.append([selected_year, 11])
        row_list.append([selected_year, 12])
        sub_columns = ['Bkdate__year', 'Bkdate__month']
        ym_list = [ dict(zip(sub_columns,row)) for row in row_list ]

    for ym in ym_list:
        try:
            get_deadline = Deadline.objects.get(year=ym['Bkdate__year'], month=ym['Bkdate__month'])
            if get_deadline.regdatetime:
                ym['close'] = 1
        except Deadline.DoesNotExist:
            pass

    return render(request,'accounting/close_list.html', {'ym_list': ym_list, 'year_range': range(this_year, 1999, -1), 'selected_year': selected_year, 'accounting_management': 'active','close_list': 'active'})

@login_required(login_url='/')
def regist_close(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    if request.method == "POST":
        ym = request.POST.get('ym')
        year = ym[:4]
        month = ym[5:]
        
        main_acct = Account.objects.get(business=business, main=True)
        try:
            tblbank_last_jango = TBLBANK.objects.filter(Bkacctno=main_acct.account_number, Bkdate__year=year, Bkdate__month=month).order_by('Bkdate','id').last().Bkjango
            transaction_last_jango = Transaction.objects.filter(business=business, Bkdate__year=year, Bkdate__month=month).order_by('Bkdate', 'id').last().Bkjango
        except TBLBANK.DoesNotExist:
            return HttpResponse("<script>alert('등록된 거래가 없습니다.');history.back();</script>")
        except AttributeError:
            return HttpResponse("<script>alert('등록된 거래가 없습니다.');history.back();</script>")

        if tblbank_last_jango == transaction_last_jango:
            deadline, created = Deadline.objects.get_or_create(business=business, year=year, month=month)
            deadline.regdatetime = today
            deadline.save()
        else:
            return HttpResponse("<script>alert('주계좌의 잔액과 거래내역의 잔액이 일치하지 않습니다.');history.back();</script>")
        
    response = redirect('close_list')
    if business.type3_id == "어린이집" and int(month) < 3:
        response['Location'] += '?year='+str(int(year)-1) #회기년도
    else:
        response['Location'] += '?year='+year   #회기년도
    return response

@login_required(login_url='/')
def undo_close(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    if request.method == "POST":
        ym = request.POST.get('ym')
        year = ym[:4]
        month = ym[5:]

        try:
            deadline = Deadline.objects.get(business=business,year=year, month=month)
            deadline.regdatetime = None
            deadline.save()
        except:
            pass

    response = redirect('close_list')
    if business.type3_id == "어린이집" and int(month) < 3:
        response['Location'] += '?year='+str(int(year)-1) #회기년도
    else:
        response['Location'] += '?year='+year   #회기년도
    return response

#--------------파일다운로드-------------
from .models import UploadFile

@login_required(login_url='/')
def file_download(request):
    upload_files = UploadFile.objects.filter(user__is_staff=True)

    return render(request, 'accounting/file_download.html', {'download': 'active', 'upload_files': upload_files})

@login_required(login_url='/')
def popup_upload(request):
    type = request.GET.get('type','')
    return render(request, 'accounting/popup_upload.html', {'type': type})

import openpyxl
from openpyxl import Workbook, load_workbook

@login_required(login_url='/')
def upload_transaction(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    ymd = DateFormat(today).format("ymdHis")
    upfile = request.FILES.get('file')
    upload_type = request.POST.get('type')

    if upfile == None:
        return HttpResponse("<script>alert('파일을 선택해주세요.');history.back();</script>")
    upfile.name = str(ymd)+'_'+business.name+'_거래내역.xlsx'
    UploadFile.objects.create(title=str(ymd)+'_'+business.name+'_거래내역', file=upfile, user=request.user)

    wb = load_workbook(filename='./media/'+upfile.name)
    sheet = wb.worksheets[0]
    try:
        Bkacct = Account.objects.get(account_number=sheet['B1'].value)
    except:
        return HttpResponse("<script>alert('등록되지 않은 계좌입니다. 엑셀파일의 계좌번호를 확인해주세요.');history.back();</script>")

    num = TBLBANK.objects.all().order_by('-Bkid').first().Bkid + 1
    for index in range(5, sheet.max_row + 1):
        if sheet['A'+str(index)] and sheet['B'+str(index)] and sheet['E'+str(index)] and (sheet['C'+str(index)] or sheet['D'+str(index)]):
            Bkinput = 0
            Bkoutput = 0
            if sheet['C'+str(index)].value and sheet['C'+str(index)].value > 0:
                Bkinput = sheet['C'+str(index)].value
            elif sheet['D'+str(index)].value and sheet['D'+str(index)].value > 0:
                Bkoutput = sheet['D'+str(index)].value

            datetimecell = sheet['A'+str(index)].value
            if type(datetimecell) == datetime.datetime:
                Bkdate = datetimecell
            else:
                try:
                    datetimecell = datetimecell.replace(".","-").replace("/","-")
                    Bkdate = datetime.datetime.strptime(datetimecell, '%Y-%m-%d %H:%M:%S')
                except:
                    created_file.delete()
                    return HttpResponse("<script>alert('"+str(index-5)+"개의 거래 등록 완료.<br>"+str(index)+"행의 날짜형식이 맞지 않습니다.');history.back();</script>")

            TBLBANK.objects.create(
                Bkid=num,
                Bkdivision=1,
                Mid=request.user.username,
                Bkacctno=Bkacct.account_number,
                Bkname=Bkacct.bank.name,
                Bkdate=Bkdate,
                Bkjukyo=sheet['B'+str(index)].value,
                Bkinput=Bkinput,
                Bkoutput=Bkoutput,
                Bkjango=sheet['E'+str(index)].value,
                business=business
            )
        num += 1

    return HttpResponse('<script type="text/javascript">window.close(); window.opener.parent.location.reload(); window.parent.location.href="/";</script>')

@login_required(login_url='/')
def upload_transaction2(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    ymd = DateFormat(today).format("ymdHis")
    upfile = request.FILES.get('file')
    upload_type = request.POST.get('type')

    if upfile == None:
        return HttpResponse("<script>alert('파일을 선택해주세요.');history.back();</script>")
    upfile.name = str(ymd)+'_'+business.name+'_거래내역.xlsx'
    created_file = UploadFile.objects.create(title=str(ymd)+'_'+business.name+'_거래내역', file=upfile, user=request.user)

    wb = load_workbook(filename='./media/'+upfile.name)
    sheet = wb.worksheets[0]
    try:
        Bkacct = Account.objects.get(account_number=sheet['B1'].value)
    except:
        created_file.delete()
        return HttpResponse("<script>alert('등록되지 않은 계좌입니다. 엑셀파일의 계좌번호를 확인해주세요.');history.back();</script>")

    num = TBLBANK.objects.all().order_by('-Bkid').first().Bkid + 1
    for index in range(5, sheet.max_row + 1):
        if sheet['A'+str(index)] and sheet['B'+str(index)] and sheet['C'+str(index)] and sheet['F'+str(index)] and (sheet['D'+str(index)] or sheet['E'+str(index)]):
            Bkinput = 0
            Bkoutput = 0
            if sheet['D'+str(index)].value and sheet['D'+str(index)].value > 0:
                Bkinput = sheet['D'+str(index)].value
            elif sheet['E'+str(index)].value and sheet['E'+str(index)].value > 0:
                Bkoutput = sheet['E'+str(index)].value

            datecell = sheet['A'+str(index)].value
            if type(datecell) == datetime.datetime:
                date = DateFormat(sheet['A'+str(index)].value).format("Y-m-d ")
            else:
                try:
                    datecell = datecell.replace(".","-").replace("/","-")
                    date = datecell+" "
                except:
                    created_file.delete()
                    return HttpResponse("<script>alert('"+str(index-5)+"개의 거래 등록 완료.<br>"+str(index)+"행의 날짜형식이 맞지 않습니다.');history.back();</script>")

            timecell = sheet['B'+str(index)].value
            if type(timecell) == datetime.datetime:
                time = DateFormat(sheet['B'+str(index)].value).format("H:i:s")
            else:
                try:
                    time = timecell
                except:
                    created_file.delete()
                    return HttpResponse("<script>alert('"+str(index-5)+"개의 거래 등록 완료.<br>"+str(index)+"행의 시간형식이 맞지 않습니다.');history.back();</script>")

            try:
                Bkdate = datetime.datetime.strptime(date+time, '%Y-%m-%d %H:%M:%S')
            except:
                created_file.delete()
                return HttpResponse("<script>alert('"+str(index-5)+"개의 거래 등록 완료.<br>"+str(index)+"행의 날짜 혹은 시간형식이 맞지 않습니다.');history.back();</script>")

            TBLBANK.objects.create(
                Bkid=num,
                Bkdivision=1,
                Mid=request.user.username,
                Bkacctno=Bkacct.account_number,
                Bkname=Bkacct.bank.name,
                Bkdate=Bkdate,
                Bkjukyo=sheet['C'+str(index)].value,
                Bkinput=Bkinput,
                Bkoutput=Bkoutput,
                Bkjango=sheet['F'+str(index)].value,
                business=business
            )
        num += 1

    return HttpResponse('<script type="text/javascript">window.close(); window.opener.parent.location.reload(); window.parent.location.href="/";</script>')

@login_required(login_url='/')
def test(request):
    return render(request,'accounting/test.html')

@login_required(login_url='/')
def design_test(request):
    return render(request,'accounting/design_test.html')

def check_date(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    year = int(request.POST.get('year'))
    month = int(request.POST.get('month'))
    year2 = int(request.POST.get('year2'))
    month2 = int(request.POST.get('month2'))
    error_message = None

    if business.type3_id == "어린이집" :
        if year > year2:
            year2 = year
            month2 = month
        elif year == year2:
            if month < 3:
                if month2 >= 3:
                    month2 = 2
                elif month > month2:
                    month2 = month
            else:
                if month > month2:
                    month2 = month
                if month2 < 3:
                    month2 = 12
        else:
            if year2 - year > 1:
                year2 = year+1
            if month < 3:
                year2 = year
                month2 = month
            else:
                if month2 > 2:
                    month2 = 2
    else:
        if year > year2:
            year2 = year
            if month > month2:
                month2 = month
        elif year == year2:
            if month > month2:
                month2 = month
        else:
            year2 = year
            if month > month2:
                month2 = month

    context = {'year': year, 'month': month, 'year2': year2, 'month2': month2, 'error_message': error_message}
    return HttpResponse(json.dumps(context), content_type="application/json")


def readNumber(strNum):
    # 만 단위 자릿수
    tenThousandPos = 4
    # 억 단위 자릿수
    hundredMillionPos = 9
    txtDigit = ['', '십', '백', '천', '만', '억']
    txtNumber = ['', '일', '이', '삼', '사', '오', '육', '칠', '팔', '구']
    txtPoint = '쩜 '
    resultStr = ''
    #자릿수 카운트
    digitCount = len(strNum) - 1
    
    index = 0
    while index < len(strNum):
        showDigit = True
        ch = strNum[index]
        
        #----------숫자표기----------
        #자릿수가 2자리이상이고 1이면 '일'은 표시 안함.
        # 단 '만' '억'에서는 표시 함
        if(digitCount > 1) and (digitCount != tenThousandPos) and  (digitCount != hundredMillionPos) and int(ch) == 1:
            if index == 0 :
                resultStr = resultStr + txtNumber[int(ch)]
            else:
                resultStr = resultStr + ''
        #0이면 숫자표기 안함
        elif int(ch) == 0:
            resultStr = resultStr + ''
            #단 '만,'억'에서는 단위표시해줌
            if (digitCount != tenThousandPos) and  (digitCount != hundredMillionPos):
                showDigit = False
        else:
            resultStr = resultStr + txtNumber[int(ch)]
                
        #----------단위표기----------
        # 1억 이상
        if digitCount > hundredMillionPos:
            if showDigit:
                resultStr = resultStr + txtDigit[digitCount-hundredMillionPos]
        # 1만 이상
        elif digitCount > tenThousandPos:
            if showDigit:
                resultStr = resultStr + txtDigit[digitCount-tenThousandPos]
        else:
            if showDigit:
                resultStr = resultStr + txtDigit[digitCount]
                
        digitCount = digitCount - 1
        index = index + 1

    return resultStr


@login_required
def set_proofnum(request):
    ym_list = Transaction.objects.exclude(Bkdivision=0).values('business','item__paragraph__subsection__type','Bkdate__year','Bkdate__month').distinct()
    #print(ym_list)
    for ym in ym_list:
        index = 1
        tr_list = Transaction.objects.filter(business=ym['business'], item__paragraph__subsection__type=ym['item__paragraph__subsection__type'], Bkdate__year=ym['Bkdate__year'], Bkdate__month=ym['Bkdate__month']).exclude(Bkdivision=0).order_by('id')
        for tr in tr_list:
            #print(ym['business'], index, tr)
            tr.proofnum = index
            tr.save()
            index += 1
        #print(index-1)

    return render(request, 'accounting/other_settings.html')


'''
@login_required
def premonth_transfer_price(request):
    Bkid = request.POST.get('Bkid')
    Bkinfo = Account.objects.get(id=Bkid)
    business = Business.objects.get(id=1)
    today = datetime.datetime.now()
    this_month = DateFormat(today).format("Y-m")+'-01'
    one_month_ago = DateFormat(today - relativedelta(months=1)).format("Y-m")+'-01'


    if request.method == "POST":
        data = TBLBANK.objects.filter(Bkacctno=Bkinfo.account_number.replace('-','')).filter(Bkdate__gte=one_month_ago).filter(Bkdate__lt=this_month).order_by('-Bkdate','-Bkid').first()
    
    is_regist = True

    try :
        transaction = Transaction.objects.filter(Bkacctno=Bkinfo.account_number).filter(Bkdate__gte=this_month).get(Bkdivision=0)
    except IndexError:
        return HttpResponse("<script>alert('해당계좌의 거래내역이 없습니다.');history.back();</script>")
    except Transaction.DoesNotExist:
        transaction = Transaction(
            Bkid=data.Bkid,
            Bkdivision=0,
            Mid=request.user.username,
            Bkacctno=Bkinfo.account_number,
            Bkname=Bkinfo.bank.name,
            Bkdate=this_month,
            Bkjukyo="전월이월금",
            Bkinput=data.Bkjango,
            Bkoutput=0,
            Bkjango=data.Bkjango,
            regdatetime=today,
        )
        transaction.save()
        is_regist = False

    if is_regist:
        return HttpResponse("<script>alert('해당계좌의 전월이월금은 입력된 상태입니다. 계좌를 확인하고 다시 등록해주세요.');history.back();</script>")
    
    return redirect('transaction_history', business.pk)
    #return HttpResponse(json.dumps(response), content_type='application/json')
'''
